"""
Final integration test for the bill summary sheet addition feature
"""
import pandas as pd
from core.processors.excel_processor import ExcelProcessor
from core.processors.batch_processor import add_bill_summary_sheet
import shutil
from pathlib import Path

def test_bill_processing_with_summary_sheet():
    """Test the complete bill processing with summary sheet addition"""
    print("=== BILL PROCESSING WITH SUMMARY SHEET TEST ===\n")
    
    # Process a test file
    processor = ExcelProcessor()
    with open('TEST_INPUT_FILES/0511-N-extra.xlsx', 'rb') as f:
        processed_data = processor.process_excel(f)
    
    print("1. Processed Data Overview:")
    print(f"   - Title data entries: {len(processed_data.get('title_data', {}))}")
    print(f"   - Work order rows: {len(processed_data.get('work_order_data', []))}")
    print(f"   - Bill quantity rows: {len(processed_data.get('bill_quantity_data', []))}")
    
    # Show key title data
    title_data = processed_data.get('title_data', {})
    print(f"\n2. Key Title Data:")
    print(f"   - Contractor: {title_data.get('Name of Contractor or supplier :', 'N/A')}")
    print(f"   - Agreement No: {title_data.get('Agreement No.', 'N/A')}")
    print(f"   - Bill Number: {title_data.get('Bill Number', 'N/A')}")
    
    # Test sheet name generation
    from core.processors.batch_processor import generate_sheet_name
    sheet_name = generate_sheet_name(title_data)
    print(f"\n3. Generated Sheet Name: {sheet_name}")
    
    # Create test file and add summary sheet
    test_output_dir = Path('FINAL_TEST_OUTPUT')
    test_output_dir.mkdir(exist_ok=True)
    test_file = test_output_dir / '0511-N-extra_with_summary.xlsx'
    
    # Copy original file
    shutil.copy2('TEST_INPUT_FILES/0511-N-extra.xlsx', test_file)
    
    print(f"\n4. Adding Summary Sheet:")
    print(f"   - Target file: {test_file}")
    result = add_bill_summary_sheet(str(test_file), processed_data)
    print(f"   - Result: {'SUCCESS' if result else 'FAILED'}")
    
    # Verify the result
    if result:
        print(f"\n5. Verification:")
        try:
            from openpyxl import load_workbook
            wb = load_workbook(test_file)
            print(f"   - All sheets: {wb.sheetnames}")
            
            if sheet_name in wb.sheetnames:
                print(f"   - ✓ Summary sheet '{sheet_name}' found")
                ws = wb[sheet_name]
                
                # Show sheet content
                print(f"   - Sheet content:")
                for i, row in enumerate(ws.iter_rows(values_only=True), 1):
                    print(f"     Row {i}: {row}")
                    if i >= 10:  # Limit output
                        print("     ...")
                        break
            else:
                print(f"   - ✗ Summary sheet '{sheet_name}' NOT found")
                
        except Exception as e:
            print(f"   - Error verifying result: {e}")
    
    print(f"\n6. Test Complete!")
    print(f"   Output file: {test_file.absolute()}")

if __name__ == "__main__":
    test_bill_processing_with_summary_sheet()