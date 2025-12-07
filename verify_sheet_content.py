import pandas as pd
from openpyxl import load_workbook
import os

# Check the content of the newly added sheet
file_path = "TEST_INPUT_FILES/0511-N-extra.xlsx"

if os.path.exists(file_path):
    # Load workbook
    wb = load_workbook(file_path, keep_vba=True)
    
    # Check the newly added sheet
    sheet_name = "Shree 48"
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"Content of sheet '{sheet_name}':")
        print(f"Dimensions: {ws.max_row} rows, {ws.max_column} columns")
        
        # Print first 20 rows
        print("\nFirst 20 rows of the sheet:")
        for row in range(1, min(21, ws.max_row + 1)):
            row_data = []
            for col in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=row, column=col).value
                row_data.append(str(cell_value) if cell_value is not None else "")
            print(f"Row {row}: {', '.join(row_data)}")
    else:
        print(f"Sheet '{sheet_name}' not found")
        
    # Also check the original sheets to see what was copied
    print("\nOriginal sheets for comparison:")
    for sheet_name in wb.sheetnames:
        if sheet_name != "Shree 48":
            ws = wb[sheet_name]
            print(f"  {sheet_name}: {ws.max_row} rows, {ws.max_column} columns")
else:
    print(f"File {file_path} not found")