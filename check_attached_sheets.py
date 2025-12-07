import pandas as pd
from openpyxl import load_workbook
import os

# Check the sheets in one of the Excel files in ATTACHED_ASSETS
file_path = "ATTACHED_ASSETS/even BILL NOTE SHEET.xlsm"

if os.path.exists(file_path):
    # Load workbook
    wb = load_workbook(file_path, keep_vba=True)
    print("Sheets in the workbook:")
    for sheet_name in wb.sheetnames:
        print(f"  - {sheet_name}")
    
    # Check if it's a macro-enabled file
    print(f"\nFile type: {'Macro-enabled' if wb.vba_archive else 'Regular'}")
    
    # Show some info about each sheet
    print("\nSheet details:")
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"  {sheet_name}: {ws.max_row} rows, {ws.max_column} columns")
        
    # Check if there's a sheet that might be the template for the new sheet
    print("\nLooking for template sheet...")
    for sheet_name in wb.sheetnames:
        if "note" in sheet_name.lower() or "bill" in sheet_name.lower():
            print(f"  Found potential template sheet: {sheet_name}")
else:
    print(f"File {file_path} not found")
    
    # List all files in ATTACHED_ASSETS to see what's available
    print("\nFiles in ATTACHED_ASSETS:")
    for file in os.listdir("ATTACHED_ASSETS"):
        if file.endswith(".xlsm"):
            print(f"  - {file}")