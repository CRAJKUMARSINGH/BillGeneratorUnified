import pandas as pd
from openpyxl import load_workbook
import os

# Check specific cells in the newly added sheet
file_path = "TEST_INPUT_FILES/0511-N-extra.xlsx"

if os.path.exists(file_path):
    # Load workbook
    wb = load_workbook(file_path, keep_vba=True)
    
    # Check the newly added sheet
    sheet_name = "Shree 48"
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"Checking specific cells in sheet '{sheet_name}':")
        
        # Check the cells we tried to update
        cells_to_check = ['A1', 'A2', 'A3', 'A4', 'A5', 'A6', 'A7', 'A8']
        for cell in cells_to_check:
            cell_obj = ws[cell]
            print(f"Cell {cell}: Value='{cell_obj.value}', Data_type='{cell_obj.data_type}'")
            
        # Check if there are formulas in these cells
        for cell in cells_to_check:
            cell_obj = ws[cell]
            if cell_obj.data_type == 'f':  # formula
                print(f"Cell {cell} contains formula: {cell_obj.value}")
    else:
        print(f"Sheet '{sheet_name}' not found")