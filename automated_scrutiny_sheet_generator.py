#!/usr/bin/env python3
"""
Automated Scrutiny Sheet Generator for Bill Processing
===============================================

This script automates the creation of scrutiny sheets for processed bills:
1. Copies template sheet within the same workbook
2. Renames sheet with contractor name + agreement number
3. Populates required data cells
4. Runs macro programmatically
5. Exports sheet as PDF

Requirements:
- Windows OS with Excel installed
- pywin32 library (pip install pywin32)
- openpyxl library (pip install openpyxl)
"""

import os
import sys
from pathlib import Path
from datetime import datetime
import re

# Global flags for library availability
WIN32COM_AVAILABLE = False
OPENPYXL_AVAILABLE = False
PANDAS_AVAILABLE = False

# Import libraries with error handling
try:
    import win32com.client
    WIN32COM_AVAILABLE = True
except ImportError:
    win32com = None
    print("WARNING: win32com not available. Install with: pip install pywin32")

try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    openpyxl = None
    print("WARNING: openpyxl not available. Install with: pip install openpyxl")

try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    pd = None
    print("WARNING: pandas not available. Install with: pip install pandas")


def get_first_n_words(text, n=5):
    """
    Extract first n words from text
    
    Args:
        text (str): Input text
        n (int): Number of words to extract
        
    Returns:
        str: First n words joined by spaces
    """
    if not text or not isinstance(text, str):
        return ""
    
    # Clean and split text
    words = re.findall(r'\b\w+\b', text.strip())
    return " ".join(words[:n]) if words else ""


def sanitize_sheet_name(name):
    """
    Sanitize sheet name to comply with Excel restrictions
    
    Args:
        name (str): Sheet name to sanitize
        
    Returns:
        str: Sanitized sheet name (max 31 chars, no invalid chars)
    """
    if not name:
        return "ScrutinySheet"
    
    # Remove invalid characters for Excel sheet names
    invalid_chars = [':', '\\', '/', '?', '*', '[', ']']
    for char in invalid_chars:
        name = name.replace(char, '_')
    
    # Remove any remaining control characters
    name = ''.join(c for c in name if ord(c) >= 32)
    
    # Truncate to 31 characters (Excel limit)
    if len(name) > 31:
        name = name[:31].rstrip()
    
    return name


def generate_sheet_name(title_data):
    """
    Generate sheet name: first 5 words of contractor + "_" + agreement number
    
    Args:
        title_data (dict): Dictionary containing title sheet data
        
    Returns:
        str: Generated sheet name or None if data insufficient
    """
    # Look for contractor name variations
    contractor_keys = [
        "Name of Contractor or supplier :",
        "Name of Contractor",
        "Contractor Name",
        "Contractor",
        "Supplier Name"
    ]
    
    # Look for agreement number variations
    agreement_keys = [
        "Agreement No.",
        "Agreement No",
        "Agreement Number",
        "Agreement",
        "Work Order No",
        "Work Order Number"
    ]
    
    contractor_name = ""
    agreement_number = ""
    
    # Find contractor name
    for key in contractor_keys:
        if key in title_data and title_data[key]:
            contractor_name = str(title_data[key]).strip()
            break
    
    # Find agreement number
    for key in agreement_keys:
        if key in title_data and title_data[key]:
            agreement_number = str(title_data[key]).strip()
            break
    
    if not contractor_name or not agreement_number:
        print(f"Warning: Missing contractor ({bool(contractor_name)}) or agreement ({bool(agreement_number)}) data")
        return None
    
    # Get first 5 words of contractor name
    contractor_prefix = get_first_n_words(contractor_name, 5)
    
    # Combine with agreement number
    sheet_name = f"{contractor_prefix}_{agreement_number}"
    
    # Sanitize for Excel
    sheet_name = sanitize_sheet_name(sheet_name)
    
    return sheet_name


def populate_scrutiny_sheet_data(ws, processed_data, bill_type="running"):
    """
    Populate scrutiny sheet cells with data from processed bill
    
    Required cells:
    - C3: Agreement Number
    - C8: Name of Work
    - C9: Name of Contractor
    - C12: Date of Commencement
    - C13: Date of Completion
    - C14: "WIP" for running bills, else Actual Completion Date
    - C17: Work Order Amount
    - C18: Last Bill Amount (0 for first/final bills)
    - C19: This Bill Amount
    - C29: Sum of Extra Items
    
    Args:
        ws: Excel worksheet object
        processed_data (dict): Processed bill data
        bill_type (str): "running", "first", or "final"
    """
    title_data = processed_data.get('title_data', {})
    totals = processed_data.get('totals', {})
    
    def get_value_from_variations(data_dict, keys, default=""):
        """Helper to get value with multiple key variations"""
        for key in keys:
            if key in data_dict and data_dict[key]:
                return data_dict[key]
        return default
    
    # Collect all the values we need to set
    cell_values = {}
    
    # C3: Agreement Number
    agreement_number = get_value_from_variations(
        title_data,
        ["Agreement No.", "Agreement No", "Agreement Number", "Work Order No", "Work Order Number"],
        ""
    )
    if agreement_number:
        cell_values['C3'] = agreement_number
    
    # C8: Name of Work
    work_name = get_value_from_variations(
        title_data,
        ["Name of Work ;-)", "Name of Work", "Project Name", "Work Description", "Name of Work ;-", "Work Name"],
        ""
    )
    if work_name:
        cell_values['C8'] = work_name
    
    # C9: Name of Contractor
    contractor_name = get_value_from_variations(
        title_data,
        ["Name of Contractor or supplier :", "Name of Contractor", "Contractor Name", "Supplier Name", "Contractor"],
        ""
    )
    if contractor_name:
        cell_values['C9'] = contractor_name
    
    # C12: Date of Commencement
    commencement_date = get_value_from_variations(
        title_data,
        ["Date of Commencement", "Start Date", "Commencement Date", "St. date of Start :", "Start Date :"],
        ""
    )
    if commencement_date:
        cell_values['C12'] = commencement_date
    
    # C13: Date of Completion
    completion_date = get_value_from_variations(
        title_data,
        ["Date of Completion", "Schedule Completion", "Completion Date", "St. date of completion :", "Completion Date :"],
        ""
    )
    if completion_date:
        cell_values['C13'] = completion_date
    
    # C14: WIP for running bills, else actual completion date
    if bill_type.lower() == "running":
        cell_values['C14'] = "WIP"  # WORK IN PROGRESS
    else:
        actual_completion = get_value_from_variations(
            title_data,
            ["Date of actual completion of work :", "Actual Date of Completion", "Actual Completion", "Actual Completion Date"],
            ""
        )
        cell_values['C14'] = actual_completion if actual_completion else completion_date
    
    # C17: Work Order Amount
    work_order_amount = totals.get('work_order_amount', 0)
    if not work_order_amount:
        work_order_amount_str = get_value_from_variations(
            title_data,
            ["WORK ORDER AMOUNT RS.", "Work Order Amount", "Work Order Value", "Work Order Amount Rs."],
            "0"
        )
        try:
            work_order_amount = float(work_order_amount_str) if work_order_amount_str else 0
        except (ValueError, TypeError):
            work_order_amount = 0
    cell_values['C17'] = work_order_amount
    
    # C18: Last Bill Amount (0 for first/final bills)
    if bill_type.lower() in ["first", "final"]:
        cell_values['C18'] = 0
    else:
        last_bill_amount = totals.get('last_bill_amount', 0)
        # If not available in totals, try to get from title data
        if not last_bill_amount:
            last_bill_amount_str = get_value_from_variations(
                title_data,
                ["Amount Paid Vide Last Bill", "Last Bill Amount", "Previous Bill Amount"],
                "0"
            )
            try:
                last_bill_amount = float(last_bill_amount_str) if last_bill_amount_str else 0
            except (ValueError, TypeError):
                last_bill_amount = 0
        cell_values['C18'] = last_bill_amount if last_bill_amount else 0
    
    # C19: This Bill Amount
    this_bill_amount = totals.get('net_payable', totals.get('grand_total', 0))
    if not this_bill_amount:
        this_bill_amount = totals.get('this_bill_amount', 0)
    if not this_bill_amount:
        this_bill_amount_str = get_value_from_variations(
            title_data,
            ["This Bill Amount", "Current Bill Amount"],
            "0"
        )
        try:
            this_bill_amount = float(this_bill_amount_str) if this_bill_amount_str else 0
        except (ValueError, TypeError):
            this_bill_amount = 0
    cell_values['C19'] = this_bill_amount
    
    # C29: Sum of Extra Items including Tender Premium
    extra_items_sum = totals.get('extra_items_sum', 0)
    tender_premium = totals.get('tender_premium', 0)
    
    # If not in totals, try to get from title data
    if not extra_items_sum:
        extra_items_sum_str = get_value_from_variations(
            title_data,
            ["Extra Items Amount", "Sum of Extra Items"],
            "0"
        )
        try:
            extra_items_sum = float(extra_items_sum_str) if extra_items_sum_str else 0
        except (ValueError, TypeError):
            extra_items_sum = 0
    
    if not tender_premium:
        tender_premium_str = get_value_from_variations(
            title_data,
            ["Tender Premium", "Premium Amount"],
            "0"
        )
        try:
            tender_premium = float(tender_premium_str) if tender_premium_str else 0
        except (ValueError, TypeError):
            tender_premium = 0
    
    total_extra = float(extra_items_sum) + float(tender_premium)
    cell_values['C29'] = total_extra
    
    # Now set all the values using the worksheet object
    for cell_address, value in cell_values.items():
        try:
            ws[cell_address].value = value
        except Exception as e:
            print(f"Warning: Could not set value for cell {cell_address}: {e}")
    
    return True


def copy_template_sheet_within_workbook(workbook_path, template_sheet_name, new_sheet_name):
    """
    Copy template sheet within the same workbook using COM automation
    
    Args:
        workbook_path (str): Path to Excel workbook (.xlsm)
        template_sheet_name (str): Name of template sheet to copy
        new_sheet_name (str): Name for new copied sheet
        
    Returns:
        bool: True if successful, False otherwise
    """
    if not WIN32COM_AVAILABLE:
        print("ERROR: win32com not available. Cannot copy sheet with macros.")
        return False
    
    excel = None
    wb = None
    try:
        # Create Excel application
        if WIN32COM_AVAILABLE and win32com:
            excel = win32com.client.Dispatch("Excel.Application")
            # Check if Excel was created successfully
            if excel is None:
                print("ERROR: Could not create Excel application object")
                return False
                
            # Test if Excel is responsive
            try:
                version = excel.Version
                print(f"Excel version: {version}")
            except Exception as e:
                print(f"ERROR: Excel application not responsive: {e}")
                return False
        else:
            print("ERROR: Could not create Excel application")
            return False
            
        excel.Visible = False
        excel.DisplayAlerts = False
        
        # Open workbook
        wb = excel.Workbooks.Open(os.path.abspath(workbook_path))
        
        # Find template sheet
        template_sheet = None
        for sheet in wb.Worksheets:
            if sheet.Name == template_sheet_name:
                template_sheet = sheet
                break
        
        if not template_sheet:
            print(f"ERROR: Template sheet '{template_sheet_name}' not found")
            wb.Close(SaveChanges=False)
            excel.Quit()
            return False
        
        # Copy template sheet
        template_sheet.Copy(After=wb.Worksheets(wb.Worksheets.Count))
        
        # Get the newly created sheet (should be the last one)
        new_sheet = wb.Worksheets(wb.Worksheets.Count)
        
        # Rename the new sheet
        new_sheet.Name = new_sheet_name
        
        # Save and close
        wb.Save()
        wb.Close(SaveChanges=True)
        excel.Quit()
        
        print(f"Successfully copied sheet '{template_sheet_name}' to '{new_sheet_name}'")
        return True
        
    except Exception as e:
        print(f"ERROR copying sheet: {e}")
        try:
            if wb:
                wb.Close(SaveChanges=False)
        except:
            pass
        try:
            if excel:
                excel.Quit()
        except:
            pass
        return False


def run_macro_in_sheet(workbook_path, sheet_name, macro_cell="E37"):
    """
    Run macro associated with a button in specified cell
    
    Args:
        workbook_path (str): Path to Excel workbook
        sheet_name (str): Name of sheet containing macro button
        macro_cell (str): Cell reference containing macro button (default: E37)
        
    Returns:
        bool: True if macro executed, False otherwise
    """
    if not WIN32COM_AVAILABLE:
        print("WARNING: win32com not available. Cannot run macro programmatically.")
        return False
    
    excel = None
    wb = None
    try:
        # Create Excel application
        if WIN32COM_AVAILABLE and win32com:
            excel = win32com.client.Dispatch("Excel.Application")
            # Check if Excel was created successfully
            if excel is None:
                print("ERROR: Could not create Excel application object")
                return False
                
            # Test if Excel is responsive
            try:
                version = excel.Version
            except Exception as e:
                print(f"ERROR: Excel application not responsive: {e}")
                return False
        else:
            print("ERROR: Could not create Excel application")
            return False
            
        excel.Visible = False
        excel.DisplayAlerts = False
        
        # Open workbook
        wb = excel.Workbooks.Open(os.path.abspath(workbook_path))
        
        # Activate the sheet
        ws = wb.Worksheets(sheet_name)
        ws.Activate()
        
        # Try different methods to run the macro
        macro_executed = False
        
        # Method 1: Try to find and click the shape/button in the specified cell
        try:
            # Iterate through all shapes to find one in the macro cell
            shapes = ws.Shapes
            for i in range(1, shapes.Count + 1):
                try:
                    shape = shapes(i)
                    # Check if shape is in the macro cell
                    if hasattr(shape, 'TopLeftCell') and shape.TopLeftCell.Address == f"${macro_cell[0]}${macro_cell[1:]}" or \
                       hasattr(shape, 'Placement'):
                        # Try to execute the shape's action
                        try:
                            # Try to run the macro directly by common names
                            excel.Application.Run("GenerateBillNotes")
                            macro_executed = True
                            print("Executed macro: GenerateBillNotes")
                            break
                        except:
                            try:
                                excel.Application.Run(f"{wb.Name}!GenerateBillNotes")
                                macro_executed = True
                                print(f"Executed macro: {wb.Name}!GenerateBillNotes")
                                break
                            except:
                                pass
                except:
                    continue
        except:
            pass
        
        # Method 2: Try to run common macro names directly
        if not macro_executed:
            common_macros = ["GenerateBillNotes", "GenerateNotes", "CreateNotes"]
            for macro_name in common_macros:
                try:
                    excel.Application.Run(macro_name)
                    macro_executed = True
                    print(f"Executed macro: {macro_name}")
                    break
                except:
                    try:
                        excel.Application.Run(f"'{wb.Name}'!{macro_name}")
                        macro_executed = True
                        print(f"Executed macro: '{wb.Name}'!{macro_name}")
                        break
                    except:
                        pass
        
        # Save and close
        wb.Save()
        wb.Close(SaveChanges=True)
        excel.Quit()
        
        if not macro_executed:
            print("WARNING: Could not execute macro automatically. Please run manually.")
            return False
            
        print("Macro executed successfully")
        return True
        
    except Exception as e:
        print(f"ERROR running macro: {e}")
        try:
            if wb:
                wb.Close(SaveChanges=False)
        except:
            pass
        try:
            if excel:
                excel.Quit()
        except:
            pass
        return False


def export_sheet_to_pdf(workbook_path, sheet_name, output_pdf_path):
    """
    Export specific sheet to PDF
    
    Args:
        workbook_path (str): Path to Excel workbook
        sheet_name (str): Name of sheet to export
        output_pdf_path (str): Path for output PDF file
        
    Returns:
        bool: True if successful, False otherwise
    """
    if not WIN32COM_AVAILABLE:
        print("WARNING: win32com not available. Cannot export to PDF.")
        return False
    
    excel = None
    wb = None
    try:
        # Create Excel application
        if WIN32COM_AVAILABLE and win32com:
            excel = win32com.client.Dispatch("Excel.Application")
            # Check if Excel was created successfully
            if excel is None:
                print("ERROR: Could not create Excel application object")
                return False
                
            # Test if Excel is responsive
            try:
                version = excel.Version
            except Exception as e:
                print(f"ERROR: Excel application not responsive: {e}")
                return False
        else:
            print("ERROR: Could not create Excel application")
            return False
            
        excel.Visible = False
        excel.DisplayAlerts = False
        
        # Open workbook
        wb = excel.Workbooks.Open(os.path.abspath(workbook_path))
        
        # Activate the sheet
        ws = wb.Worksheets(sheet_name)
        ws.Activate()
        
        # Export to PDF
        ws.ExportAsFixedFormat(
            Type=0,  # xlTypePDF
            Filename=os.path.abspath(output_pdf_path),
            Quality=0,  # xlQualityStandard
            IncludeDocProperties=True,
            IgnorePrintAreas=False,
            OpenAfterPublish=False
        )
        
        # Close
        wb.Close(SaveChanges=False)
        excel.Quit()
        
        print(f"Successfully exported sheet '{sheet_name}' to PDF: {output_pdf_path}")
        return True
        
    except Exception as e:
        print(f"ERROR exporting to PDF: {e}")
        try:
            if wb:
                wb.Close(SaveChanges=False)
        except:
            pass
        try:
            if excel:
                excel.Quit()
        except:
            pass
        return False


def calculate_totals_from_data(processed_data):
    """
    Calculate totals from processed_data if not already present
    
    Args:
        processed_data (dict): Processed bill data
        
    Returns:
        dict: Totals dictionary
    """
    # If totals already exist, return them
    if 'totals' in processed_data and processed_data['totals']:
        return processed_data['totals']
    
    totals = {}
    title_data = processed_data.get('title_data', {})
    
    # Get work order and extra items data
    work_order_data = processed_data.get('work_order_data', None)
    extra_items_data = processed_data.get('extra_items_data', None)
    
    # Create empty DataFrame if pandas is available, otherwise use None
    if not PANDAS_AVAILABLE or pd is None:
        work_order_data = work_order_data if work_order_data is not None else []
        extra_items_data = extra_items_data if extra_items_data is not None else []
    else:
        if work_order_data is None:
            work_order_data = pd.DataFrame()
        if extra_items_data is None:
            extra_items_data = pd.DataFrame()
    
    # Calculate work order amount
    work_order_amount = 0
    if PANDAS_AVAILABLE and pd is not None and isinstance(work_order_data, pd.DataFrame):
        if not work_order_data.empty:
            for index, row in work_order_data.iterrows():
                try:
                    qty = float(row.get('Quantity Since', row.get('Quantity', 0)) or 0)
                    rate = float(row.get('Rate', 0) or 0)
                    work_order_amount += qty * rate
                except (ValueError, TypeError):
                    pass
    elif isinstance(work_order_data, (list, dict)):
        # Handle list or dict format
        for item in work_order_data if isinstance(work_order_data, list) else [work_order_data]:
            try:
                qty = float(item.get('Quantity Since', item.get('Quantity', 0)) or 0)
                rate = float(item.get('Rate', 0) or 0)
                work_order_amount += qty * rate
            except (ValueError, TypeError):
                pass
    
    # Calculate extra items sum
    extra_items_sum = 0
    if PANDAS_AVAILABLE and pd is not None and isinstance(extra_items_data, pd.DataFrame):
        if not extra_items_data.empty:
            for index, row in extra_items_data.iterrows():
                try:
                    qty = float(row.get('Quantity', 0) or 0)
                    rate = float(row.get('Rate', 0) or 0)
                    extra_items_sum += qty * rate
                except (ValueError, TypeError):
                    pass
    elif isinstance(extra_items_data, (list, dict)):
        # Handle list or dict format
        for item in extra_items_data if isinstance(extra_items_data, list) else [extra_items_data]:
            try:
                qty = float(item.get('Quantity', 0) or 0)
                rate = float(item.get('Rate', 0) or 0)
                extra_items_sum += qty * rate
            except (ValueError, TypeError):
                pass
    
    # Get tender premium
    tender_premium_percent = 0
    try:
        tender_premium_percent = float(title_data.get('TENDER PREMIUM %', 0) or 0)
    except (ValueError, TypeError):
        pass
    
    tender_premium_amount = work_order_amount * (tender_premium_percent / 100)
    
    # Get last bill amount
    last_bill_amount = 0
    try:
        last_bill_amount_str = title_data.get('Amount Paid Vide Last Bill', title_data.get('Last Bill Amount', 0))
        last_bill_amount = float(last_bill_amount_str or 0)
    except (ValueError, TypeError):
        pass
    
    # Calculate net payable (this bill amount)
    grand_total = work_order_amount + tender_premium_amount
    net_payable = grand_total - last_bill_amount
    
    totals = {
        'work_order_amount': work_order_amount,
        'extra_items_sum': extra_items_sum,
        'tender_premium': tender_premium_amount,
        'tender_premium_amount': tender_premium_amount,
        'last_bill_amount': last_bill_amount,
        'net_payable': net_payable,
        'grand_total': grand_total
    }
    
    return totals


def create_scrutiny_sheet_for_bill(
    workbook_path, 
    processed_data, 
    bill_type="running",
    template_sheet_name=None,
    output_pdf_dir=None
):
    """
    Main function to create scrutiny sheet for a processed bill
    
    For each new bill process:
    - Add a sheet by copying the template sheet within the same workbook
    - Rename the sheet by first 5 words of contractor + "_" + agreement number
    - Populate data in C3, C8, C9, C12-C14, C17-C19, C29 from bill processed
    - For running bills, C14 shall be "WIP" (Work In Progress)
    - For first running or final bill, C18 is zero, else from "Amount Paid Vide Last Bill"
    - C29 is sum of extra items including tender premium (can be zero)
    - After filling, programmatically run the macro button in cell E37
    - Print the sheet as "MACRO scrutiny SHEET IN PDF"
    
    Args:
        workbook_path (str): Path to Excel workbook (.xlsm) - the template file
        processed_data (dict): Processed bill data containing title_data and optionally totals
        bill_type (str): Type of bill - "running", "first", or "final"
        template_sheet_name (str): Name of template sheet (default: first sheet)
        output_pdf_dir (str): Directory for PDF output (default: same as workbook)
        
    Returns:
        dict: Result with success status and details
    """
    workbook_path = Path(workbook_path)
    
    if not workbook_path.exists():
        return {
            'success': False,
            'error': f"Workbook not found: {workbook_path}"
        }
    
    if not WIN32COM_AVAILABLE:
        return {
            'success': False,
            'error': "Required library 'win32com' not available. Install with: pip install pywin32"
        }
    
    try:
        # Calculate totals if not provided
        if 'totals' not in processed_data or not processed_data.get('totals'):
            processed_data['totals'] = calculate_totals_from_data(processed_data)
        
        # Generate sheet name
        title_data = processed_data.get('title_data', {})
        sheet_name = generate_sheet_name(title_data)
        
        if not sheet_name:
            return {
                'success': False,
                'error': "Could not generate sheet name - missing contractor or agreement data"
            }
        
        print(f"Generated sheet name: {sheet_name}")
        
        # Use first sheet as template if not specified
        if not template_sheet_name:
            try:
                if OPENPYXL_AVAILABLE and openpyxl:
                    wb_temp = openpyxl.load_workbook(str(workbook_path), keep_vba=True)
                    template_sheet_name = wb_temp.sheetnames[0] if wb_temp.sheetnames else "Sheet1"
                    wb_temp.close()
                else:
                    template_sheet_name = "Sheet1"
            except:
                template_sheet_name = "Sheet1"
        
        print(f"Using template sheet: {template_sheet_name}")
        
        # Copy template sheet within workbook
        if not copy_template_sheet_within_workbook(str(workbook_path), template_sheet_name, sheet_name):
            return {
                'success': False,
                'error': "Failed to copy template sheet"
            }
        
        # Load workbook and populate data using openpyxl first (more reliable for cell access)
        try:
            if OPENPYXL_AVAILABLE and openpyxl:
                # Load workbook with openpyxl
                wb = openpyxl.load_workbook(str(workbook_path), keep_vba=True)
                ws = wb[sheet_name]
                
                # Populate data
                print("Populating sheet data...")
                populate_scrutiny_sheet_data(ws, processed_data, bill_type)
                
                # Save workbook
                wb.save(str(workbook_path))
                wb.close()
                print("Data populated successfully using openpyxl")
            else:
                print("WARNING: openpyxl not available, falling back to COM for data population")
                # Fallback to COM if openpyxl is not available
                excel = None
                wb = None
                if WIN32COM_AVAILABLE and win32com:
                    excel = win32com.client.Dispatch("Excel.Application")
                    # Check if Excel was created successfully
                    if excel is None:
                        print("ERROR: Could not create Excel application object")
                        return {
                            'success': False,
                            'error': "Could not create Excel application"
                        }
                        
                    # Test if Excel is responsive
                    try:
                        version = excel.Version
                    except Exception as e:
                        print(f"ERROR: Excel application not responsive: {e}")
                        return {
                            'success': False,
                            'error': "Excel application not responsive"
                        }
                else:
                    print("ERROR: Could not create Excel application")
                    return {
                        'success': False,
                        'error': "Could not create Excel application"
                    }
                    
                excel.Visible = False
                excel.DisplayAlerts = False
                
                wb = excel.Workbooks.Open(os.path.abspath(workbook_path))
                ws = wb.Worksheets(sheet_name)
                
                # Populate data
                print("Populating sheet data...")
                populate_scrutiny_sheet_data(ws, processed_data, bill_type)
                
                # Save and close
                wb.Save()
                wb.Close(SaveChanges=True)
                excel.Quit()
        except Exception as e:
            print(f"ERROR populating data: {e}")
            return {
                'success': False,
                'error': f"Failed to populate sheet data: {e}"
            }
        
        # Run macro
        print("Running macro...")
        macro_success = run_macro_in_sheet(str(workbook_path), sheet_name)
        
        # Export to PDF
        pdf_path = None
        if output_pdf_dir:
            output_pdf_dir = Path(output_pdf_dir)
            output_pdf_dir.mkdir(parents=True, exist_ok=True)
            # Use the exact format requested: "MACRO scrutiny SHEET IN PDF"
            pdf_filename = f"MACRO scrutiny SHEET IN PDF_{sheet_name}.pdf"
            pdf_path = output_pdf_dir / pdf_filename
        else:
            # Default to same directory as workbook
            # Use the exact format requested: "MACRO scrutiny SHEET IN PDF"
            pdf_path = workbook_path.parent / f"MACRO scrutiny SHEET IN PDF_{sheet_name}.pdf"
        
        print("Exporting to PDF...")
        pdf_success = export_sheet_to_pdf(str(workbook_path), sheet_name, str(pdf_path))
        
        return {
            'success': True,
            'sheet_name': sheet_name,
            'workbook_path': str(workbook_path),
            'pdf_path': str(pdf_path) if pdf_success else None,
            'macro_executed': macro_success,
            'pdf_exported': pdf_success
        }
        
    except Exception as e:
        return {
            'success': False,
            'error': str(e)
        }


# Example usage
if __name__ == "__main__":
    print("Automated Scrutiny Sheet Generator")
    print("==================================")
    
    # Example data structure (this would come from your bill processing)
    example_processed_data = {
        'title_data': {
            'Name of Contractor or supplier :': 'M/s. Shree Krishna Builders Jaipur',
            'Agreement No.': '48/2024-25',
            'Name of Work ;-)': 'Plumbing Installation and MTC work at Govt. Nehru hostel Mansarovar, Sanganer, Jaipur',
            'Date of Commencement': '2024-01-15',
            'Date of Completion': '2024-12-15',
            'Date of actual completion of work :': '2024-12-10',
            'WORK ORDER AMOUNT RS.': 500000
        },
        'totals': {
            'work_order_amount': 500000,
            'last_bill_amount': 200000,
            'net_payable': 250000,
            'extra_items_sum': 15000,
            'tender_premium': 0
        }
    }
    
    # Example usage
    workbook_path = "ATTACHED_ASSETS/even BILL NOTE SHEET.xlsm"
    
    if os.path.exists(workbook_path):
        result = create_scrutiny_sheet_for_bill(
            workbook_path=workbook_path,
            processed_data=example_processed_data,
            bill_type="running",
            output_pdf_dir="output_pdfs"
        )
        
        if result['success']:
            print(f"\n✅ SUCCESS!")
            print(f"   Sheet Name: {result['sheet_name']}")
            print(f"   Workbook: {result['workbook_path']}")
            print(f"   PDF Exported: {result['pdf_exported']}")
            if result['pdf_path']:
                print(f"   PDF Path: {result['pdf_path']}")
            print(f"   Macro Executed: {result['macro_executed']}")
        else:
            print(f"\n❌ FAILED: {result['error']}")
    else:
        print(f"\n❌ Workbook not found: {workbook_path}")
        print("Please ensure the template workbook exists in the ATTACHED_ASSETS folder.")