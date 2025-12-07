import pandas as pd
import re
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

def extract_contractor_first_name(contractor_name):
    """Extract the first name from contractor name, skipping prefixes like 'M/s.'"""
    if not contractor_name:
        return ""
    
    # Remove common prefixes
    prefixes = ['M/s.', 'M/S.', 'Mr.', 'Mrs.', 'Ms.']
    cleaned_name = contractor_name.strip()
    
    for prefix in prefixes:
        if cleaned_name.startswith(prefix):
            cleaned_name = cleaned_name[len(prefix):].strip()
            break
    
    # Extract first word/name
    words = cleaned_name.split()
    if words:
        # Return first 5 letters of the first word
        return words[0][:5]
    return ""

def extract_agreement_number_without_year(agreement_number):
    """Extract agreement number without year part"""
    if not agreement_number:
        return ""
    
    # Handle format like "48/2024-25" -> we want "48"
    parts = str(agreement_number).split('/')
    if len(parts) > 1:
        # Take the part before the slash
        return parts[0]
    else:
        # If no slash, just return the whole thing
        return str(agreement_number)

def generate_sheet_name(title_data):
    """Generate sheet name based on contractor name and agreement number"""
    contractor_key = "Name of Contractor or supplier :"
    agreement_key = "Agreement No."
    
    contractor_name = title_data.get(contractor_key, "")
    agreement_number = title_data.get(agreement_key, "")
    
    if not contractor_name or not agreement_number:
        return None
    
    first_5_letters = extract_contractor_first_name(contractor_name)
    agreement_without_year = extract_agreement_number_without_year(agreement_number)
    
    if first_5_letters and agreement_without_year:
        return f"{first_5_letters} {agreement_without_year}"
    
    return None

def create_bill_summary_data(processed_data):
    """Create a summary dataframe with key bill information"""
    title_data = processed_data.get('title_data', {})
    work_order_data = processed_data.get('work_order_data', pd.DataFrame())
    
    # Extract key information
    summary_data = {
        'Field': [],
        'Value': []
    }
    
    # Add title information
    key_mappings = {
        'Bill Number': 'Bill Number',
        'Name of Contractor or supplier :': 'Contractor',
        'Agreement No.': 'Agreement No.',
        'Name of Work ;-': 'Work Description',
        'WORK ORDER AMOUNT RS.': 'Work Order Amount',
        'St. date of Start :': 'Start Date',
        'St. date of completion :': 'Completion Date',
        'Date of actual completion of work :': 'Actual Completion Date'
    }
    
    for display_name, key_name in key_mappings.items():
        value = title_data.get(key_name, '')
        if value:
            summary_data['Field'].append(display_name)
            summary_data['Value'].append(str(value))
    
    # Add financial summary
    if not work_order_data.empty:
        total_amount = 0
        for index, row in work_order_data.iterrows():
            quantity = pd.to_numeric(row.get('Quantity', 0), errors='coerce')
            rate = pd.to_numeric(row.get('Rate', 0), errors='coerce')
            amount = quantity * rate
            total_amount += amount
        
        summary_data['Field'].append('Calculated Total Amount')
        summary_data['Value'].append(f"{total_amount:.2f}")
    
    return pd.DataFrame(summary_data)

def add_bill_summary_sheet(file_path, processed_data):
    """
    Add a new sheet to the Excel file with bill summary information
    
    Args:
        file_path: Path to the Excel file
        processed_data: Processed data from ExcelProcessor
    """
    try:
        # Generate sheet name
        sheet_name = generate_sheet_name(processed_data.get('title_data', {}))
        if not sheet_name:
            print("Could not generate sheet name - missing contractor or agreement data")
            return False
        
        # Create summary data
        summary_df = create_bill_summary_data(processed_data)
        
        # Load the workbook
        wb = load_workbook(file_path)
        
        # Check if sheet already exists, if so, remove it
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
        
        # Create new sheet
        ws = wb.create_sheet(sheet_name)
        
        # Add data to the sheet
        for r in dataframe_to_rows(summary_df, index=False, header=True):
            ws.append(r)
        
        # Save the workbook
        wb.save(file_path)
        print(f"Successfully added sheet '{sheet_name}' to {file_path}")
        return True
        
    except Exception as e:
        print(f"Error adding bill summary sheet: {e}")
        return False

# Test the function
if __name__ == "__main__":
    # This would typically be called from the batch processor
    print("Bill summary sheet addition module ready")