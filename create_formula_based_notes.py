"""
Create formula-based note generation at A40
The note will be generated using Excel formulas that concatenate text with cell references
"""
import openpyxl
import shutil
from datetime import datetime

source_file = 'ATTACHED_ASSETS/english Note FINAL BILL NOTE SHEET.xlsm'
backup_file = f'ATTACHED_ASSETS/english Note FINAL BILL NOTE SHEET_backup_formulanotes_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsm'

print("="*70)
print("CREATING FORMULA-BASED NOTE GENERATION")
print("="*70)

# Create backup
shutil.copy2(source_file, backup_file)
print(f"\n✓ Backup created: {backup_file}")

# Load the file
wb = openpyxl.load_workbook(source_file, keep_vba=True)
ws = wb.active

print(f"✓ Loaded workbook")

# First, add computing helper cells (hidden area)
print("\n✓ Adding computing helper formulas (E34-E39):")

# E34: Percentage of work done
ws.cell(34, 5).value = "=IF(C18,ROUND((C21/C18)*100,2),0)"
print(f"  E34: Percentage work done")

# E35: Time allowed (days)
ws.cell(35, 5).value = "=IF(AND(C14,C13),C14-C13,0)"
print(f"  E35: Time allowed days")

# E36: Delay days
ws.cell(36, 5).value = "=IF(AND(C15,C14),C15-C14,0)"
print(f"  E36: Delay days")

# E37: Extra item percentage
ws.cell(37, 5).value = "=IF(AND(C31,C18),ROUND((C31/C18)*100,2),0)"
print(f"  E37: Extra item %")

# E38: Is delay > 50% of time?
ws.cell(38, 5).value = "=IF(E35>0,IF(E36>(E35*0.5),\"SE\",\"Office\"),\"\")"
print(f"  E38: Delay approval level")

# E39: Deviation approval level based on percentage
ws.cell(39, 5).value = "=IF(E34<90,\"<90\",IF(E34>105,\"SE\",IF(E34>100,\"Office\",\"\")))"
print(f"  E39: Deviation approval level")

# Now create the formula-based note at A40
print("\n✓ Creating formula-based note at A40:")

# Build the note using TEXTJOIN or CONCATENATE
# Note: Excel 2016+ has TEXTJOIN, older versions need CONCATENATE

note_formula = '''=
"1. The work has been completed " & TEXT(E34,"0.00") & "% of the Work Order Amount." & CHAR(10) &
IF(E34<90,"2. The execution of work at final stage is less than 90% of the Work Order Amount, the Requisite Deviation Statement is enclosed to observe check on unuseful expenditure. Approval of the Deviation is having jurisdiction under this office." & CHAR(10),"") &
IF(AND(E34>100,E34<=105),"2. Requisite Deviation Statement is enclosed. The Overall Excess is less than or equal to 5% and is having approval jurisdiction under this office." & CHAR(10),"") &
IF(E34>105,"2. Requisite Deviation Statement is enclosed. The Overall Excess is more than 5% and Approval of the Deviation Case is required from the Superintending Engineer, PWD Electrical Circle, Udaipur." & CHAR(10),"") &
IF(E36>0,"3. Time allowed for completion of the work was " & E35 & " days. The Work was delayed by " & E36 & " days." & CHAR(10) & "4. Approval of the Time Extension Case is required from the " & IF(E38="SE","Superintending Engineer, PWD Electrical Circle, Udaipur.","this office.") & CHAR(10),"3. Work was completed in time." & CHAR(10)) &
IF(C30="Yes",IF(E37>5,"5. The amount of Extra items is Rs. " & C31 & " which is " & TEXT(E37,"0.00") & "% of the Work Order Amount; exceed 5%, require approval from the Superintending Engineer, PWD Electrical Circle, Udaipur." & CHAR(10),"5. The amount of Extra items is Rs. " & C31 & " which is " & TEXT(E37,"0.00") & "% of the Work Order Amount; under 5%, approval of the same is to be granted by this office." & CHAR(10)),"") &
IF(C32="Yes",IF(E34<=100,"6. The details of the items wherein EXCESS QUANTITY has been executed during the work completion are attached. There is a saving in the work of the tune of " & TEXT(100-E34,"0.00") & "%; (i.e., Overall Excess = NIL, just excess quantity items are recorded), the approval of which falls under the jurisdiction of this office." & CHAR(10),"6. The details of the items wherein EXCESS QUANTITY has been executed during the work completion are attached. There is an Overall excess of " & TEXT(E34-100,"0.00") & "% requires approval as already narrated." & CHAR(10)),"") &
"7. Quality Control (QC) test reports attached." & CHAR(10) &
IF(C29="No","8. A copy of Hand Over statement duly signed by client department representative is attached." & CHAR(10),"") &
IF(C33="Yes","9. The Bill is submitted very late; explanation from concerned Assistant Engineer may be sought." & CHAR(10),"") &
"10. Please peruse above details for necessary decision-making." & CHAR(10) & CHAR(10) &
"                      Premlata Jain" & CHAR(10) &
"                     AAO- As Auditor"
'''

# Clean up the formula (remove line breaks and extra spaces)
note_formula = note_formula.replace('\n', '').replace('  ', ' ')

ws.cell(40, 1).value = note_formula
print(f"  A40: Formula-based note created")

# Set row height and formatting
ws.row_dimensions[40].height = 315
ws.cell(40, 1).alignment = openpyxl.styles.Alignment(wrap_text=True, vertical='top', horizontal='left')

# Merge A40:D40
ws.merge_cells('A40:D40')
print(f"  A40:D40 merged, height=315")

# Save the file
wb.save(source_file)
print(f"\n✓ Saved file with formula-based notes")

print("\n" + "="*70)
print("FORMULA-BASED NOTE GENERATION CREATED")
print("="*70)
print("\n✓ Computing formulas in E34-E39:")
print("  E34 = Percentage work done")
print("  E35 = Time allowed (days)")
print("  E36 = Delay days")
print("  E37 = Extra item %")
print("  E38 = Approval level (SE/Office)")
print("  E39 = Deviation level")
print("\n✓ Note formula at A40 (merged A40:D40):")
print("  • Automatically generates notes based on data")
print("  • References cells: C18, C19, C20, C21, C29-C33")
print("  • Uses helper formulas E34-E39")
print("  • Updates automatically when data changes")
print("\n✓ No VBA macro needed for note generation!")
print(f"✓ Backup saved: {backup_file}")
print("\n📝 Usage:")
print("  1. Fill data in C19, C20, C29-C33")
print("  2. Note automatically appears in A40")
print("  3. Print using Ctrl+P")
