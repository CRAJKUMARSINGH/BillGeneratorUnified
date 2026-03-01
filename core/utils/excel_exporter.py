"""
Excel Exporter - Round-trip Excel export with formatting preservation
Implements Phase 1.3 requirement: Excel upload → edit → re-download with formatting
"""
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import io
from typing import Optional, Dict, Any
from datetime import datetime


class ExcelExporter:
    """Export edited data back to Excel while preserving original formatting"""
    
    @staticmethod
    def export_with_formatting(
        original_file: Any,
        edited_df: pd.DataFrame,
        sheet_name: str = 'Bill Quantity',
        preserve_formulas: bool = True
    ) -> io.BytesIO:
        """
        Export edited data back to Excel while preserving original formatting
        
        Args:
            original_file: Original uploaded Excel file (BytesIO or file path)
            edited_df: Edited DataFrame with changes
            sheet_name: Sheet name to update (default: 'Bill Quantity')
            preserve_formulas: Whether to preserve formulas (default: True)
            
        Returns:
            BytesIO object containing the updated Excel file
        """
        # Load original workbook
        if isinstance(original_file, io.BytesIO):
            original_file.seek(0)
            wb = load_workbook(original_file)
        else:
            wb = load_workbook(original_file)
        
        # Get the sheet to update
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            # Create new sheet if it doesn't exist
            ws = wb.create_sheet(sheet_name)
        
        # Store original formatting (first data row as template)
        original_formats = {}
        if ws.max_row > 1:  # Has data beyond header
            for col_idx, cell in enumerate(ws[2], start=1):  # Row 2 is first data row
                original_formats[col_idx] = {
                    'font': cell.font.copy() if cell.font else None,
                    'fill': cell.fill.copy() if cell.fill else None,
                    'border': cell.border.copy() if cell.border else None,
                    'alignment': cell.alignment.copy() if cell.alignment else None,
                    'number_format': cell.number_format
                }
        
        # Find header row (usually row 1)
        header_row = 1
        
        # Map DataFrame columns to Excel columns
        column_mapping = {}
        for col_idx, cell in enumerate(ws[header_row], start=1):
            if cell.value:
                header_name = str(cell.value).strip()
                # Try to find matching column in DataFrame
                for df_col in edited_df.columns:
                    if header_name.lower() in df_col.lower() or df_col.lower() in header_name.lower():
                        column_mapping[df_col] = col_idx
                        break
        
        # Update data rows (starting from row 2)
        data_start_row = header_row + 1
        
        # Clear existing data (but keep formulas if preserve_formulas=True)
        if not preserve_formulas:
            for row_idx in range(data_start_row, ws.max_row + 1):
                for col_idx in range(1, ws.max_column + 1):
                    ws.cell(row_idx, col_idx).value = None
        
        # Write edited data
        for df_row_idx, row_data in edited_df.iterrows():
            excel_row_idx = data_start_row + df_row_idx
            
            for df_col, excel_col_idx in column_mapping.items():
                cell = ws.cell(excel_row_idx, excel_col_idx)
                
                # Update value
                value = row_data[df_col]
                
                # Handle different data types
                if pd.isna(value):
                    cell.value = None
                elif isinstance(value, (int, float)):
                    cell.value = float(value)
                else:
                    cell.value = str(value)
                
                # Apply original formatting if available
                if excel_col_idx in original_formats:
                    fmt = original_formats[excel_col_idx]
                    if fmt['font']:
                        cell.font = fmt['font']
                    if fmt['fill']:
                        cell.fill = fmt['fill']
                    if fmt['border']:
                        cell.border = fmt['border']
                    if fmt['alignment']:
                        cell.alignment = fmt['alignment']
                    if fmt['number_format']:
                        cell.number_format = fmt['number_format']
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output
    
    @staticmethod
    def create_new_excel(
        edited_df: pd.DataFrame,
        title_data: Optional[Dict[str, Any]] = None,
        include_formatting: bool = True
    ) -> io.BytesIO:
        """
        Create a new Excel file from edited data with professional formatting
        
        Args:
            edited_df: Edited DataFrame
            title_data: Optional title sheet data
            include_formatting: Whether to apply formatting
            
        Returns:
            BytesIO object containing the new Excel file
        """
        from openpyxl import Workbook
        
        wb = Workbook()
        
        # Create Title sheet if data provided
        if title_data:
            ws_title = wb.active
            ws_title.title = 'Title'
            
            # Write title data
            row_idx = 1
            for key, value in title_data.items():
                ws_title.cell(row_idx, 1).value = key
                ws_title.cell(row_idx, 2).value = value
                row_idx += 1
            
            # Create Bill Quantity sheet
            ws = wb.create_sheet('Bill Quantity')
        else:
            ws = wb.active
            ws.title = 'Bill Quantity'
        
        # Write DataFrame to sheet
        for r_idx, row in enumerate(dataframe_to_rows(edited_df, index=False, header=True), start=1):
            for c_idx, value in enumerate(row, start=1):
                cell = ws.cell(r_idx, c_idx)
                cell.value = value
                
                # Apply formatting
                if include_formatting:
                    if r_idx == 1:  # Header row
                        cell.font = Font(bold=True, color="FFFFFF")
                        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                        cell.border = Border(
                            left=Side(style='thin'),
                            right=Side(style='thin'),
                            top=Side(style='thin'),
                            bottom=Side(style='thin')
                        )
                    else:  # Data rows
                        cell.alignment = Alignment(horizontal="left", vertical="center")
                        cell.border = Border(
                            left=Side(style='thin'),
                            right=Side(style='thin'),
                            top=Side(style='thin'),
                            bottom=Side(style='thin')
                        )
                        
                        # Format numbers
                        if isinstance(value, (int, float)) and not pd.isna(value):
                            if 'Rate' in edited_df.columns[c_idx-1] or 'Amount' in edited_df.columns[c_idx-1]:
                                cell.number_format = '₹#,##0.00'
                            elif 'Quantity' in edited_df.columns[c_idx-1]:
                                cell.number_format = '#,##0.00'
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Cap at 50
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output
    
    @staticmethod
    def add_change_log_sheet(
        excel_file: io.BytesIO,
        change_log_df: pd.DataFrame
    ) -> io.BytesIO:
        """
        Add a change log sheet to existing Excel file
        
        Args:
            excel_file: Excel file as BytesIO
            change_log_df: Change log DataFrame
            
        Returns:
            Updated Excel file with change log sheet
        """
        excel_file.seek(0)
        wb = load_workbook(excel_file)
        
        # Create or get Change Log sheet
        if 'Change Log' in wb.sheetnames:
            ws = wb['Change Log']
            # Clear existing data
            for row in ws.iter_rows():
                for cell in row:
                    cell.value = None
        else:
            ws = wb.create_sheet('Change Log')
        
        # Write change log data
        for r_idx, row in enumerate(dataframe_to_rows(change_log_df, index=False, header=True), start=1):
            for c_idx, value in enumerate(row, start=1):
                cell = ws.cell(r_idx, c_idx)
                cell.value = value
                
                # Format header
                if r_idx == 1:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                
                # Add borders
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 40)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output
