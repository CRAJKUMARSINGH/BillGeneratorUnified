"""
Batch Processor - Process multiple Excel files
"""
import os
from pathlib import Path
from typing import List, Dict
import streamlit as st
from concurrent.futures import ThreadPoolExecutor, as_completed
import time
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

def extract_contractor_first_name(contractor_name):
    """Extract the first name from contractor name, skipping prefixes like 'M/s.'"""
    if not contractor_name:
        return ""
    
    # Remove common prefixes
    prefixes = ['M/s.', 'M/S.', 'Mr.', 'Mrs.', 'Ms.']
    cleaned_name = contractor_name.strip()
    
    for prefix in prefixes:
        if cleaned_name.startswith(prefix):
            cleaned_name = cleaned_name[len(prefix):].strip()
            break
    
    # Extract first word/name
    words = cleaned_name.split()
    if words:
        # Return first 5 letters of the first word
        return words[0][:5]
    return ""

def extract_agreement_number_without_year(agreement_number):
    """Extract agreement number without year part"""
    if not agreement_number:
        return ""
    
    # Handle format like "48/2024-25" -> we want "48"
    parts = str(agreement_number).split('/')
    if len(parts) > 1:
        # Take the part before the slash
        return parts[0]
    else:
        # If no slash, just return the whole thing
        return str(agreement_number)

def generate_sheet_name(title_data):
    """Generate sheet name based on contractor name and agreement number"""
    contractor_key = "Name of Contractor or supplier :"
    agreement_key = "Agreement No."
    
    contractor_name = title_data.get(contractor_key, "")
    agreement_number = title_data.get(agreement_key, "")
    
    if not contractor_name or not agreement_number:
        return None
    
    first_5_letters = extract_contractor_first_name(contractor_name)
    agreement_without_year = extract_agreement_number_without_year(agreement_number)
    
    if first_5_letters and agreement_without_year:
        return f"{first_5_letters} {agreement_without_year}"
    
    return None

def create_bill_summary_data(processed_data):
    """Create a summary dataframe with key bill information"""
    title_data = processed_data.get('title_data', {})
    work_order_data = processed_data.get('work_order_data', pd.DataFrame())
    
    # Extract key information
    summary_data = {
        'Field': [],
        'Value': []
    }
    
    # Add title information
    key_mappings = {
        'Bill Number': 'Bill Number',
        'Name of Contractor or supplier :': 'Contractor',
        'Agreement No.': 'Agreement No.',
        'Name of Work ;-': 'Work Description',
        'WORK ORDER AMOUNT RS.': 'Work Order Amount',
        'St. date of Start :': 'Start Date',
        'St. date of completion :': 'Completion Date',
        'Date of actual completion of work :': 'Actual Completion Date'
    }
    
    for display_name, key_name in key_mappings.items():
        value = title_data.get(key_name, '')
        if value:
            summary_data['Field'].append(display_name)
            summary_data['Value'].append(str(value))
    
    # Add financial summary
    if not work_order_data.empty:
        total_amount = 0
        for index, row in work_order_data.iterrows():
            quantity = pd.to_numeric(row.get('Quantity', 0), errors='coerce')
            rate = pd.to_numeric(row.get('Rate', 0), errors='coerce')
            amount = quantity * rate
            total_amount += amount
        
        summary_data['Field'].append('Calculated Total Amount')
        summary_data['Value'].append(f"{total_amount:.2f}")
    
    return pd.DataFrame(summary_data)

def add_bill_summary_sheet(file_path, processed_data):
    """
    Add a new sheet to the Excel file with bill summary information
    
    Args:
        file_path: Path to the Excel file
        processed_data: Processed data from ExcelProcessor
    """
    try:
        # Generate sheet name
        sheet_name = generate_sheet_name(processed_data.get('title_data', {}))
        if not sheet_name:
            print("Could not generate sheet name - missing contractor or agreement data")
            return False
        
        # Create summary data
        summary_df = create_bill_summary_data(processed_data)
        
        # Load the workbook
        wb = load_workbook(file_path)
        
        # Check if sheet already exists, if so, remove it
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
        
        # Create new sheet
        ws = wb.create_sheet(sheet_name)
        
        # Add data to the sheet
        for r in dataframe_to_rows(summary_df, index=False, header=True):
            ws.append(r)
        
        # Save the workbook
        wb.save(file_path)
        print(f"Successfully added sheet '{sheet_name}' to {file_path}")
        return True
        
    except Exception as e:
        print(f"Error adding bill summary sheet: {e}")
        return False

class BatchProcessor:
    """Process multiple files in batch"""
    
    def __init__(self, config):
        self.config = config
        self.max_workers = 4
        self.output_base_dir = Path("output")
    
    def _create_timestamped_folder(self, filename: str) -> Path:
        """Create a timestamped folder for output files"""
        # Get current timestamp
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        
        # Clean filename (remove extension)
        clean_filename = Path(filename).stem
        
        # Create folder name: YYYYMMDD_HHMMSS_filename
        folder_name = f"{timestamp}_{clean_filename}"
        
        # Create full path
        output_folder = self.output_base_dir / folder_name
        output_folder.mkdir(parents=True, exist_ok=True)
        
        return output_folder
    
    def process_batch(self, files: List, progress_callback=None) -> Dict:
        """Process multiple files"""
        results = {}
        total = len(files)
        
        for idx, file in enumerate(files, 1):
            try:
                if progress_callback:
                    progress_callback(idx, total, file.name)
                
                # Create timestamped folder for this file
                output_folder = self._create_timestamped_folder(file.name)
                
                # Process file
                result = self._process_single_file(file, output_folder)
                results[file.name] = {
                    'status': 'success',
                    'data': result,
                    'output_folder': str(output_folder)
                }
            except Exception as e:
                results[file.name] = {
                    'status': 'error',
                    'error': str(e)
                }
        
        return results
    
    def _process_single_file(self, file, output_folder: Path, use_enhanced_pdf: bool = True):
        """Process a single file and save outputs to timestamped folder"""
        from core.processors.excel_processor import ExcelProcessor
        from core.generators.document_generator import DocumentGenerator
        
        # Process Excel file
        excel_processor = ExcelProcessor()
        processed_data = excel_processor.process_excel(file)
        
        # Generate documents
        doc_generator = DocumentGenerator(processed_data)
        html_documents = doc_generator.generate_all_documents()
        
        # Generate DOC documents
        doc_documents = doc_generator.generate_doc_documents()
        
        # Generate PDFs with enhanced generator if available
        if use_enhanced_pdf:
            try:
                from core.generators.pdf_generator_enhanced import EnhancedPDFGenerator
                print("✅ Using Enhanced PDF Generator (CSS Zoom + Disable Smart Shrinking)")
                
                pdf_gen = EnhancedPDFGenerator()
                pdf_documents_bytes = pdf_gen.batch_convert(
                    html_documents,
                    zoom=1.0,
                    disable_smart_shrinking=True
                )
                
                # Convert to expected format
                pdf_documents = {f"{name}.pdf": content for name, content in pdf_documents_bytes.items()}
                
            except Exception as e:
                print(f"⚠️ Enhanced PDF Generator failed, using fallback: {e}")
                pdf_documents = doc_generator.create_pdf_documents(html_documents)
        else:
            pdf_documents = doc_generator.create_pdf_documents(html_documents)
        
        # Save HTML files
        html_folder = output_folder / "html"
        html_folder.mkdir(exist_ok=True)
        
        for doc_name, html_content in html_documents.items():
            html_file = html_folder / f"{doc_name}.html"
            html_file.write_text(html_content, encoding='utf-8')
        
        # Save PDF files
        pdf_folder = output_folder / "pdf"
        pdf_folder.mkdir(exist_ok=True)
        
        for doc_name, pdf_content in pdf_documents.items():
            pdf_file = pdf_folder / doc_name
            pdf_file.write_bytes(pdf_content)
        
        # Save DOC files
        doc_folder = output_folder / "doc"
        doc_folder.mkdir(exist_ok=True)
        
        for doc_name, doc_content in doc_documents.items():
            doc_file = doc_folder / doc_name
            doc_file.write_bytes(doc_content)
        
        # Add summary sheet to the original Excel file
        try:
            # Save a copy of the original file with the summary sheet added
            original_file_copy = output_folder / f"{Path(file.name).stem}_with_summary_sheet.xlsx"
            # Write the original file content to the new location
            if hasattr(file, 'seek'):
                file.seek(0)
                with open(original_file_copy, 'wb') as f:
                    f.write(file.read())
                # Reset file pointer again for potential future use
                file.seek(0)
            
            # Add the summary sheet
            add_bill_summary_sheet(str(original_file_copy), processed_data)
        except Exception as e:
            print(f"Warning: Could not add summary sheet to {file.name}: {e}")
        
        return {
            'html_files': list(html_documents.keys()),
            'pdf_files': list(pdf_documents.keys()),
            'doc_files': list(doc_documents.keys()),
            'output_folder': str(output_folder)
        }

def show_batch_mode(config):
    """Show batch processing UI"""
    st.markdown("## 📦 Batch Processing Mode")
    st.info("Process multiple Excel files at once")
    
    uploaded_files = st.file_uploader(
        "Upload Excel Files",
        type=['xlsx', 'xls'],
        accept_multiple_files=True
    )
    
    if uploaded_files:
        st.success(f"✅ {len(uploaded_files)} files uploaded")
        
        if st.button("🚀 Process All Files", type="primary"):
            processor = BatchProcessor(config)
            
            progress_bar = st.progress(0)
            status_text = st.empty()
            
            def update_progress(current, total, filename):
                progress = current / total
                progress_bar.progress(progress)
                status_text.text(f"Processing {current}/{total}: {filename}")
            
            with st.spinner("Processing files..."):
                results = processor.process_batch(uploaded_files, update_progress)
            
            # Celebrate success with balloons! 🎈
            success_count = sum(1 for r in results.values() if r['status'] == 'success')
            if success_count > 0:
                st.balloons()
                st.success(f"🎉 Successfully processed {success_count} file(s)!")
            
            # Show results
            st.markdown("### 📊 Results")
            
            success_count = sum(1 for r in results.values() if r['status'] == 'success')
            error_count = len(results) - success_count
            
            col1, col2, col3 = st.columns(3)
            col1.metric("Total", len(results))
            col2.metric("Success", success_count)
            col3.metric("Errors", error_count)
            
            # Detailed results with output folder information
            st.markdown("### 📁 Output Folders")
            for filename, result in results.items():
                if result['status'] == 'success':
                    with st.expander(f"✅ {filename}", expanded=True):
                        st.success(f"**Output Folder:** `{result['output_folder']}`")
                        
                        data = result['data']
                        col1, col2 = st.columns(2)
                        
                        col1, col2, col3 = st.columns(3)
                        
                        with col1:
                            st.markdown("**📄 HTML Files:**")
                            for html_file in data.get('html_files', []):
                                st.text(f"  • {html_file}.html")
                        
                        with col2:
                            st.markdown("**📕 PDF Files:**")
                            for pdf_file in data.get('pdf_files', []):
                                st.text(f"  • {pdf_file}")
                        
                        with col3:
                            st.markdown("**📝 DOC Files:**")
                            for doc_file in data.get('doc_files', []):
                                st.text(f"  • {doc_file}")
                        
                        st.info(f"📂 All files saved to: {result['output_folder']}")
                else:
                    st.error(f"❌ {filename}: {result['error']}")
            
            # Summary of all output folders
            st.markdown("---")
            st.markdown("### 📦 All Output Folders")
            output_folders = [r['output_folder'] for r in results.values() if r['status'] == 'success']
            if output_folders:
                st.code('\n'.join(output_folders), language='text')
            
            # Add ZIP download functionality for all processed files
            if output_folders:
                st.markdown("### 📦 Download All as ZIP")
                st.info("Creating ZIP archive of all generated documents...")
                
                # Create zip of all output folders
                import zipfile
                import io
                import shutil
                
                # Create temporary zip file
                zip_buffer = io.BytesIO()
                with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
                    for folder in output_folders:
                        folder_path = Path(folder)
                        if folder_path.exists():
                            # Add all files in the folder to the zip
                            for file_path in folder_path.rglob('*'):
                                if file_path.is_file():
                                    # Add file to zip with relative path
                                    arc_name = f"{folder_path.name}/{file_path.relative_to(folder_path)}"
                                    zip_file.write(file_path, arc_name)
                
                zip_buffer.seek(0)
                
                st.download_button(
                    "📥 Download All Documents as ZIP",
                    data=zip_buffer,
                    file_name="all_bill_documents.zip",
                    mime="application/zip",
                    key="batch_zip_download"
                )
