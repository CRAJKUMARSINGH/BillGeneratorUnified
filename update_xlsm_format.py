"""
Update the english Note FINAL BILL NOTE SHEET.xlsm file to accommodate
the new format with 17.A, 17.B, and 17.C rows based on EVEN BILL_SCRUTINY_SHEET.xlsx
"""
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from copy import copy
import shutil
from datetime import datetime

# Create backup
backup_file = f'ATTACHED_ASSETS/english Note FINAL BILL NOTE SHEET_backup_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsm'
shutil.copy2('ATTACHED_ASSETS/english Note FINAL BILL NOTE SHEET.xlsm', backup_file)
print(f"✓ Backup created: {backup_file}")

# Load the XLSM file with VBA macros preserved
wb = openpyxl.load_workbook('ATTACHED_ASSETS/english Note FINAL BILL NOTE SHEET.xlsm', keep_vba=True)
ws = wb.active

print(f"✓ Loaded workbook: {wb.sheetnames}")
print(f"✓ Active sheet: {ws.title}")

# Find row 19 (current item 17)
row_17_index = None
for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=30, min_col=1, max_col=1), start=1):
    cell_value = row[0].value
    if cell_value == 17:
        row_17_index = idx
        print(f"✓ Found item 17 at row {idx}")
        break

if not row_17_index:
    print("✗ Could not find item 17 in the sheet")
    exit(1)

# Insert 2 new rows after the current row 17
ws.insert_rows(row_17_index + 1, 2)
print(f"✓ Inserted 2 new rows after row {row_17_index}")

# Now update the structure:
# Row 19 (original): 17 -> 17.A "Sum of payment upto last bill"
# Row 20 (new): 17.B "Amount of this bill"
# Row 21 (new): 17.C "Actual expenditure upto this bill = (A + B)"

# Update row 19 (17.A)
ws.cell(row_17_index, 1).value = "17.A"
ws.cell(row_17_index, 2).value = "Sum of payment upto last bill Rs."
# Keep the existing value in column C or set to 0
if ws.cell(row_17_index, 3).value and isinstance(ws.cell(row_17_index, 3).value, str) and '=' in str(ws.cell(row_17_index, 3).value):
    ws.cell(row_17_index, 3).value = 0
print(f"✓ Updated row {row_17_index} to 17.A")

# Update row 20 (17.B) - new row
ws.cell(row_17_index + 1, 1).value = "B."
ws.cell(row_17_index + 1, 2).value = "Amount of this bill Rs."
ws.cell(row_17_index + 1, 3).value = 0  # Default value
# Copy formatting from row above
for col in range(1, 5):
    source_cell = ws.cell(row_17_index, col)
    target_cell = ws.cell(row_17_index + 1, col)
    if source_cell.has_style:
        target_cell.font = copy(source_cell.font)
        target_cell.border = copy(source_cell.border)
        target_cell.fill = copy(source_cell.fill)
        target_cell.number_format = copy(source_cell.number_format)
        target_cell.protection = copy(source_cell.protection)
        target_cell.alignment = copy(source_cell.alignment)
print(f"✓ Created row {row_17_index + 1} as 17.B")

# Update row 21 (17.C) - new row
ws.cell(row_17_index + 2, 1).value = "C."
ws.cell(row_17_index + 2, 2).value = "Actual expenditure upto this bill = (A + B) Rs."
ws.cell(row_17_index + 2, 3).value = f"=C{row_17_index}+C{row_17_index + 1}"
# Copy formatting
for col in range(1, 5):
    source_cell = ws.cell(row_17_index, col)
    target_cell = ws.cell(row_17_index + 2, col)
    if source_cell.has_style:
        target_cell.font = copy(source_cell.font)
        target_cell.border = copy(source_cell.border)
        target_cell.fill = copy(source_cell.fill)
        target_cell.number_format = copy(source_cell.number_format)
        target_cell.protection = copy(source_cell.protection)
        target_cell.alignment = copy(source_cell.alignment)
print(f"✓ Created row {row_17_index + 2} as 17.C")

# Update row numbers for subsequent items (18, 19, 20, etc.)
# They are now 2 rows down
for row_idx in range(row_17_index + 3, ws.max_row + 1):
    cell_value = ws.cell(row_idx, 1).value
    if isinstance(cell_value, int) and cell_value >= 18:
        # Item numbers remain the same (18, 19, 20, etc.)
        pass
    
    # Update formulas that reference row 18 (which was the old item 17)
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row_idx, col_idx)
        if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
            formula = cell.value
            # Update references: C18 should now reference C21 (17.C)
            # C19 should now reference C21 (17.C) if it was referencing old item 17
            if 'C18' in formula:
                # Old row 18 was item 16 (Amount of Work Order)
                # It's now at the same position, so no change needed
                pass
            if f'C{row_17_index}' in formula and row_idx > row_17_index + 2:
                # References to old item 17 should now reference 17.C
                new_formula = formula.replace(f'C{row_17_index}', f'C{row_17_index + 2}')
                cell.value = new_formula
                print(f"✓ Updated formula in row {row_idx}, col {col_idx}: {formula} -> {new_formula}")

# Find and update the "Balance to be done" formula (item 18)
# It should be: =C18-C21 (Amount of Work Order - Actual Expenditure)
for row_idx in range(row_17_index + 3, min(row_17_index + 10, ws.max_row + 1)):
    if ws.cell(row_idx, 1).value == 18:
        # Find the row with item 16 (Amount of Work Order)
        item_16_row = None
        for r in range(1, row_17_index):
            if ws.cell(r, 1).value == 16:
                item_16_row = r
                break
        
        if item_16_row:
            ws.cell(row_idx, 3).value = f"=C{item_16_row}-C{row_17_index + 2}"
            print(f"✓ Updated item 18 formula: =C{item_16_row}-C{row_17_index + 2}")
        break

# Save the updated file
output_file = 'ATTACHED_ASSETS/english Note FINAL BILL NOTE SHEET_UPDATED.xlsm'
wb.save(output_file)
print(f"\n✓ Successfully saved updated file: {output_file}")

# Display the updated structure
print("\n" + "="*60)
print("UPDATED STRUCTURE:")
print("="*60)
for row_idx in range(row_17_index - 2, min(row_17_index + 8, ws.max_row + 1)):
    row_data = [ws.cell(row_idx, col).value for col in range(1, 5)]
    print(f"Row {row_idx}: {row_data}")

print("\n✓ Update complete!")
print(f"✓ Original file backed up to: {backup_file}")
print(f"✓ Updated file saved to: {output_file}")
print("\nPlease review the updated file and replace the original if everything looks correct.")
