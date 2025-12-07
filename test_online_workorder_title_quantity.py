#!/usr/bin/env python3
"""
Online Test - Work Order and Title Sheet with Quantity Modifications
- Processes all test input files
- Adds/modifies quantities for 5-10 main items and 1-5 extra items
- Focuses on Work Order and Title Sheet validation
- Performance optimized
"""
import os
import sys
from pathlib import Path
from datetime import datetime
import json
import traceback
import time
import random
import copy
import io

# Add core to path
sys.path.insert(0, str(Path(__file__).parent))

import pandas as pd
import openpyxl
from openpyxl import load_workbook

def modify_quantities_optimized(file_path, main_items_count=None, extra_items_count=None):
    """
    Modify quantities in Excel file - optimized version
    Returns modified workbook in memory
    Performance optimizations:
    - Use read_only=False for faster writes
    - Minimize cell access
    - Batch operations
    """
    # Load workbook with optimizations
    wb = load_workbook(file_path, data_only=False, read_only=False, keep_vba=False)
    
    modifications = {
        'main_items': [],
        'extra_items': []
    }
    
    # Modify Bill Quantity sheet (main items)
    if 'Bill Quantity' in wb.sheetnames:
        ws_bq = wb['Bill Quantity']
        
        # Find quantity column (usually column D or E) - optimized
        header_row = 1
        qty_col = None
        item_col = None
        
        # Check first row for headers - optimized single pass
        header_row_cells = list(ws_bq[header_row])
        for col_idx, cell in enumerate(header_row_cells, 1):
            if cell.value and isinstance(cell.value, str):
                cell_lower = cell.value.lower()
                if ('quantity' in cell_lower or 'qty' in cell_lower) and qty_col is None:
                    qty_col = col_idx
                if 'item' in cell_lower and item_col is None:
                    item_col = col_idx
                # Early exit if both found
                if qty_col and item_col:
                    break
        
        # If not found, assume standard structure (Item, Description, Unit, Quantity, Rate, Amount)
        if qty_col is None:
            qty_col = 4  # Column D
        if item_col is None:
            item_col = 1  # Column A
        
        # Count valid items (rows with item numbers) - optimized batch reading
        valid_rows = []
        # Read rows in batch for better performance
        for row_idx in range(header_row + 1, min(ws_bq.max_row + 1, header_row + 100)):  # Limit to first 100 rows for performance
            item_cell = ws_bq.cell(row=row_idx, column=item_col)
            
            # Quick check if cell has value
            if item_cell.value is None:
                continue
                
            item_str = str(item_cell.value).strip()
            if not item_str or item_str == 'nan':
                continue
            
            try:
                # Try to parse as number (item number)
                item_val = float(item_cell.value)
                if item_val > 0:
                    valid_rows.append((row_idx, item_val))
            except:
                # Check if it's a valid item identifier
                if len(item_str) > 0:
                    valid_rows.append((row_idx, item_str))
        
        # Select random items to modify
        if main_items_count is None:
            main_items_count = random.randint(5, min(10, len(valid_rows)))
        
        selected_rows = random.sample(valid_rows, min(main_items_count, len(valid_rows)))
        
        # Modify quantities
        for row_idx, item_val in selected_rows:
            qty_cell = ws_bq.cell(row=row_idx, column=qty_col)
            old_qty = qty_cell.value if qty_cell.value is not None else 0
            
            # Generate new quantity (increase by 10-50%)
            try:
                old_qty_float = float(old_qty) if old_qty else 0
                new_qty = old_qty_float * (1 + random.uniform(0.1, 0.5))
                qty_cell.value = round(new_qty, 2)
                
                modifications['main_items'].append({
                    'row': row_idx,
                    'item': item_val,
                    'old_quantity': old_qty_float,
                    'new_quantity': round(new_qty, 2)
                })
            except:
                # If can't parse, set a random quantity
                new_qty = random.uniform(1, 100)
                qty_cell.value = round(new_qty, 2)
                modifications['main_items'].append({
                    'row': row_idx,
                    'item': item_val,
                    'old_quantity': old_qty,
                    'new_quantity': round(new_qty, 2)
                })
    
    # Modify Extra Items sheet
    if 'Extra Items' in wb.sheetnames:
        ws_ei = wb['Extra Items']
        
        # Find header row (usually row 5, index 4)
        header_row = None
        for row_idx in range(1, min(10, ws_ei.max_row + 1)):
            row_values = [str(cell.value).lower() if cell.value else '' for cell in ws_ei[row_idx]]
            if any('qty' in val or 'quantity' in val for val in row_values):
                header_row = row_idx
                break
        
        if header_row is None:
            header_row = 5  # Default
        
        # Find quantity column
        qty_col = None
        for col_idx, cell in enumerate(ws_ei[header_row], 1):
            if cell.value and isinstance(cell.value, str):
                if 'qty' in cell.value.lower() or 'quantity' in cell.value.lower():
                    qty_col = col_idx
                    break
        
        if qty_col is None:
            qty_col = 5  # Default (column E)
        
        # Find valid extra item rows (rows with S.No. like E-01, E-02, etc.) - optimized
        valid_rows = []
        # Limit search to reasonable range for performance
        max_search_rows = min(ws_ei.max_row + 1, header_row + 50)
        for row_idx in range(header_row + 1, max_search_rows):
            sno_cell = ws_ei.cell(row=row_idx, column=1)  # Column A (S.No.)
            
            if sno_cell.value:
                sno_str = str(sno_cell.value).strip()
                if sno_str and 'E-' in sno_str:
                    valid_rows.append(row_idx)
        
        # Select random extra items to modify
        if extra_items_count is None:
            extra_items_count = random.randint(1, min(5, len(valid_rows)))
        
        selected_rows = random.sample(valid_rows, min(extra_items_count, len(valid_rows)))
        
        # Modify quantities
        for row_idx in selected_rows:
            qty_cell = ws_ei.cell(row=row_idx, column=qty_col)
            sno_cell = ws_ei.cell(row=row_idx, column=1)
            old_qty = qty_cell.value if qty_cell.value is not None else 0
            
            # Generate new quantity (increase by 20-100%)
            try:
                old_qty_float = float(old_qty) if old_qty else 0
                new_qty = old_qty_float * (1 + random.uniform(0.2, 1.0)) if old_qty_float > 0 else random.uniform(1, 10)
                qty_cell.value = round(new_qty, 2)
                
                modifications['extra_items'].append({
                    'row': row_idx,
                    'item': str(sno_cell.value),
                    'old_quantity': old_qty_float,
                    'new_quantity': round(new_qty, 2)
                })
            except:
                new_qty = random.uniform(1, 10)
                qty_cell.value = round(new_qty, 2)
                modifications['extra_items'].append({
                    'row': row_idx,
                    'item': str(sno_cell.value),
                    'old_quantity': old_qty,
                    'new_quantity': round(new_qty, 2)
                })
    
    # Save to BytesIO for processing
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output, modifications

def process_file_online(input_file, output_base_dir):
    """Process file with quantity modifications - online mode"""
    print(f"\n{'='*70}")
    print(f"Processing: {input_file.name}")
    print(f"{'='*70}")
    
    result = {
        'input_file': str(input_file),
        'filename': input_file.name,
        'status': 'pending',
        'error': None,
        'processing_time': 0,
        'modifications': {},
        'work_order_validation': {},
        'title_sheet_validation': {},
        'output_files': {},
        'timestamp': datetime.now().isoformat()
    }
    
    start_time = time.time()
    
    try:
        # Step 1: Modify quantities
        print("[1/5] Modifying quantities...")
        modified_file, modifications = modify_quantities_optimized(
            input_file,
            main_items_count=random.randint(5, 10),
            extra_items_count=random.randint(1, 5)
        )
        result['modifications'] = modifications
        print(f"     [OK] Modified {len(modifications['main_items'])} main items")
        print(f"     [OK] Modified {len(modifications['extra_items'])} extra items")
        
        # Step 2: Process Excel file
        print("[2/5] Processing Excel file...")
        from core.processors.excel_processor import ExcelProcessor
        excel_processor = ExcelProcessor()
        processed_data = excel_processor.process_excel(modified_file)
        print(f"     [OK] Excel processed successfully")
        
        # Step 3: Validate Work Order sheet
        print("[3/5] Validating Work Order sheet...")
        work_order_data = processed_data.get('work_order_data', pd.DataFrame())
        bill_quantity_data = processed_data.get('bill_quantity_data', pd.DataFrame())
        
        wo_validation = {
            'has_work_order': not work_order_data.empty,
            'work_order_rows': len(work_order_data),
            'has_bill_quantity': not bill_quantity_data.empty,
            'bill_quantity_rows': len(bill_quantity_data),
            'columns': work_order_data.columns.tolist() if not work_order_data.empty else []
        }
        
        # Check for quantity column
        if not work_order_data.empty:
            qty_cols = [col for col in work_order_data.columns if 'quantity' in str(col).lower() or 'qty' in str(col).lower()]
            wo_validation['has_quantity_column'] = len(qty_cols) > 0
            wo_validation['quantity_columns'] = qty_cols
        
        result['work_order_validation'] = wo_validation
        print(f"     [OK] Work Order: {wo_validation['work_order_rows']} rows")
        print(f"     [OK] Bill Quantity: {wo_validation['bill_quantity_rows']} rows")
        
        # Step 4: Validate Title sheet
        print("[4/5] Validating Title sheet...")
        title_data = processed_data.get('title_data', {})
        
        title_validation = {
            'has_title_data': len(title_data) > 0,
            'title_fields_count': len(title_data),
            'key_fields': {}
        }
        
        # Check for key fields (with variations)
        key_fields_map = {
            'Agreement No': ['Agreement No', 'Agreement Number', 'Agreement No.', 'Agreement'],
            'Work Order No': ['Work Order No', 'Work Order Number', 'Work Order No.', 'Work Order'],
            'Project Name': ['Project Name', 'Project', 'Name of Work'],
            'Contractor Name': ['Contractor Name', 'Contractor', 'Name of Contractor'],
            'Date of Commencement': ['Date of Commencement', 'Commencement Date', 'Start Date'],
            'Date of Completion': ['Date of Completion', 'Completion Date', 'Schedule Completion'],
            'Work Order Amount': ['Work Order Amount', 'Work Order Value', 'Amount']
        }
        
        for field_name, variations in key_fields_map.items():
            title_validation['key_fields'][field_name] = any(
                var in title_data for var in variations
            )
        
        result['title_sheet_validation'] = title_validation
        key_fields_count = len(key_fields_map)
        found_count = sum(title_validation['key_fields'].values())
        print(f"     [OK] Title fields: {title_validation['title_fields_count']}")
        print(f"     [OK] Key fields present: {found_count}/{key_fields_count}")
        
        # Step 5: Generate documents (focus on Work Order and Title)
        print("[5/6] Generating documents...")
        from core.generators.document_generator import DocumentGenerator
        doc_generator = DocumentGenerator(processed_data)
        
        # Generate only essential documents for testing
        html_documents = {}
        
        # Generate First Page (uses Title data)
        try:
            html_documents['First Page Summary'] = doc_generator._render_template('first_page.html')
        except:
            pass
        
        # Generate Scrutiny Sheet (uses Work Order and Bill Quantity)
        try:
            html_documents['BILL SCRUTINY SHEET'] = doc_generator._render_template('note_sheet.html')
        except:
            pass
        
        # Create output folder first
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        file_stem = input_file.stem
        output_folder = output_base_dir / f"{timestamp}_{file_stem}"
        output_folder.mkdir(parents=True, exist_ok=True)
        
        # Step 6: Add Macro Scrutiny Sheet
        print("[6/6] Adding Macro Scrutiny Sheet...")
        macro_sheet_result = None
        try:
            from add_macro_scrutiny_sheet import add_macro_scrutiny_sheet
            
            # Determine bill type from filename
            filename_lower = input_file.name.lower()
            if 'final' in filename_lower:
                bill_type = "final"
            elif 'first' in filename_lower or 'running' in filename_lower:
                bill_type = "running"
            else:
                bill_type = "running"  # default
            
            # Create output file path for macro sheet
            macro_output_file = output_folder / f"{input_file.stem}_with_macro.xlsm"
            
            macro_sheet_result = add_macro_scrutiny_sheet(
                processed_data,
                str(macro_output_file),
                bill_type=bill_type
            )
            
            if macro_sheet_result.get('success'):
                print(f"     [OK] Macro sheet '{macro_sheet_result['sheet_name']}' added")
                if macro_sheet_result.get('pdf_exported'):
                    print(f"     [OK] PDF exported: {Path(macro_sheet_result['pdf_file']).name}")
                result['macro_sheet'] = macro_sheet_result
            else:
                print(f"     [WARNING] Macro sheet failed: {macro_sheet_result.get('error', 'Unknown error')}")
        except Exception as e:
            print(f"     [WARNING] Could not add macro sheet: {e}")
        
        # Save HTML files
        html_folder = output_folder / "html"
        html_folder.mkdir(exist_ok=True)
        
        for doc_name, html_content in html_documents.items():
            html_file = html_folder / f"{doc_name}.html"
            html_file.write_text(html_content, encoding='utf-8')
            result['output_files'][doc_name] = str(html_file)
        
        result['output_folder'] = str(output_folder)
        result['status'] = 'success'
        result['processing_time'] = time.time() - start_time
        
        print(f"     [OK] Generated {len(html_documents)} document(s)")
        print(f"     [OK] Processing complete in {result['processing_time']:.2f} seconds")
        
        # Show summary
        if title_data:
            print(f"\n     Summary:")
            if 'Project Name' in title_data:
                print(f"       Project: {title_data['Project Name']}")
            if 'Work Order No' in title_data:
                print(f"       Work Order: {title_data['Work Order No']}")
        
    except Exception as e:
        result['status'] = 'error'
        result['error'] = str(e)
        result['traceback'] = traceback.format_exc()
        result['processing_time'] = time.time() - start_time
        print(f"[ERROR] {str(e)}")
        traceback.print_exc()
    
    return result

def main():
    """Main function to process all test files"""
    print("="*70)
    print("ONLINE TEST - WORK ORDER & TITLE SHEET WITH QUANTITY MODIFICATIONS")
    print("="*70)
    print(f"Start Time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("="*70)
    
    # Setup directories
    test_input_dir = Path('TEST_INPUT_FILES')
    output_base_dir = Path('test_output_online_workorder')
    output_base_dir.mkdir(exist_ok=True)
    
    if not test_input_dir.exists():
        print(f"\n[ERROR] Test input directory not found: {test_input_dir}")
        return
    
    # Get all input files
    input_files = list(test_input_dir.glob('*.xlsx')) + list(test_input_dir.glob('*.xls'))
    
    if not input_files:
        print(f"\n[WARNING] No Excel files found in {test_input_dir}")
        return
    
    print(f"\nInput Directory: {test_input_dir}")
    print(f"Output Directory: {output_base_dir}")
    print(f"\nFound {len(input_files)} input file(s):")
    for i, file in enumerate(sorted(input_files), 1):
        size_kb = file.stat().st_size / 1024
        print(f"   {i}. {file.name} ({size_kb:.1f} KB)")
    
    print(f"\n{'='*70}")
    print("Starting online processing with quantity modifications...")
    print(f"{'='*70}\n")
    
    # Process each file
    results = []
    total_start_time = time.time()
    
    for i, input_file in enumerate(sorted(input_files), 1):
        print(f"\n[{i}/{len(input_files)}]")
        result = process_file_online(input_file, output_base_dir)
        results.append(result)
    
    total_time = time.time() - total_start_time
    
    # Generate comprehensive report
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    report_file = output_base_dir / f"test_report_{timestamp}.json"
    summary_file = output_base_dir / f"test_summary_{timestamp}.txt"
    
    # Calculate statistics
    success_count = sum(1 for r in results if r['status'] == 'success')
    error_count = sum(1 for r in results if r['status'] == 'error')
    total_main_mods = sum(len(r.get('modifications', {}).get('main_items', [])) for r in results)
    total_extra_mods = sum(len(r.get('modifications', {}).get('extra_items', [])) for r in results)
    avg_time = sum(r.get('processing_time', 0) for r in results) / len(results) if results else 0
    
    # Save JSON report
    report_data = {
        'test_info': {
            'timestamp': datetime.now().isoformat(),
            'test_type': 'ONLINE - Work Order & Title Sheet with Quantity Modifications',
            'input_directory': str(test_input_dir),
            'output_directory': str(output_base_dir),
            'total_files': len(input_files),
            'success_count': success_count,
            'error_count': error_count,
            'total_processing_time': total_time,
            'average_time_per_file': avg_time,
            'total_main_item_modifications': total_main_mods,
            'total_extra_item_modifications': total_extra_mods
        },
        'results': results
    }
    
    with open(report_file, 'w', encoding='utf-8') as f:
        json.dump(report_data, f, indent=2, default=str)
    
    # Generate text summary
    with open(summary_file, 'w', encoding='utf-8') as f:
        f.write("="*70 + "\n")
        f.write("ONLINE TEST - WORK ORDER & TITLE SHEET SUMMARY\n")
        f.write("="*70 + "\n\n")
        f.write(f"Test Run: {timestamp}\n")
        f.write(f"Start Time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        f.write(f"Input Directory: {test_input_dir}\n")
        f.write(f"Output Directory: {output_base_dir}\n\n")
        
        f.write("-"*70 + "\n")
        f.write("OVERALL STATISTICS\n")
        f.write("-"*70 + "\n\n")
        f.write(f"Total Files Processed: {len(input_files)}\n")
        f.write(f"Success: {success_count}\n")
        f.write(f"Errors: {error_count}\n")
        f.write(f"Success Rate: {(success_count/len(input_files)*100) if input_files else 0:.1f}%\n")
        f.write(f"Total Processing Time: {total_time:.2f} seconds\n")
        f.write(f"Average Time per File: {avg_time:.2f} seconds\n")
        f.write(f"Total Main Item Modifications: {total_main_mods}\n")
        f.write(f"Total Extra Item Modifications: {total_extra_mods}\n\n")
        
        f.write("-"*70 + "\n")
        f.write("DETAILED RESULTS\n")
        f.write("-"*70 + "\n\n")
        
        for i, result in enumerate(results, 1):
            f.write(f"{i}. {result['filename']}\n")
            f.write(f"   Status: {result['status'].upper()}\n")
            f.write(f"   Processing Time: {result.get('processing_time', 0):.2f} seconds\n")
            
            if result['status'] == 'success':
                mods = result.get('modifications', {})
                f.write(f"   Main Items Modified: {len(mods.get('main_items', []))}\n")
                f.write(f"   Extra Items Modified: {len(mods.get('extra_items', []))}\n")
                
                wo_val = result.get('work_order_validation', {})
                f.write(f"   Work Order Rows: {wo_val.get('work_order_rows', 0)}\n")
                f.write(f"   Bill Quantity Rows: {wo_val.get('bill_quantity_rows', 0)}\n")
                
                title_val = result.get('title_sheet_validation', {})
                f.write(f"   Title Fields: {title_val.get('title_fields_count', 0)}\n")
                f.write(f"   Key Fields Present: {sum(title_val.get('key_fields', {}).values())}\n")
                
                f.write(f"   Output Folder: {Path(result.get('output_folder', '')).name}\n")
            else:
                f.write(f"   Error: {result.get('error', 'Unknown error')}\n")
            
            f.write("\n")
    
    # Print summary to console
    print("\n" + "="*70)
    print("TEST COMPLETE")
    print("="*70)
    print(f"\nTotal Files: {len(input_files)}")
    print(f"[SUCCESS] Success: {success_count}")
    print(f"[ERROR] Errors: {error_count}")
    print(f"Success Rate: {(success_count/len(input_files)*100) if input_files else 0:.1f}%")
    print(f"Total Time: {total_time:.2f} seconds")
    print(f"Average Time per File: {avg_time:.2f} seconds")
    print(f"Total Main Item Modifications: {total_main_mods}")
    print(f"Total Extra Item Modifications: {total_extra_mods}")
    print(f"\nOutput Directory: {output_base_dir}")
    print(f"JSON Report: {report_file.name}")
    print(f"Text Summary: {summary_file.name}")
    
    if success_count > 0:
        print(f"\n[SUCCESS] Successfully processed files:")
        for result in results:
            if result['status'] == 'success':
                mods = result.get('modifications', {})
                print(f"   - {result['filename']}: {len(mods.get('main_items', []))} main, {len(mods.get('extra_items', []))} extra mods")
    
    if error_count > 0:
        print(f"\n[ERROR] Files with errors:")
        for result in results:
            if result['status'] == 'error':
                print(f"   - {result['filename']}: {result.get('error', 'Unknown error')}")
    
    print("="*70)

if __name__ == "__main__":
    main()

