#!/usr/bin/env python3
"""
Verification script for the Automated Scrutiny Sheet Generator implementation
"""

import os
import openpyxl
from automated_scrutiny_sheet_generator import create_scrutiny_sheet_for_bill

def verify_implementation():
    """Verify that our implementation works correctly"""
    
    print("Verifying Automated Scrutiny Sheet Generator Implementation")
    print("=" * 60)
    
    # Test data
    test_data = {
        'title_data': {
            'Name of Contractor or supplier :': 'Test Contractor Ltd',
            'Agreement No.': '789/2024-25',
            'Name of Work ;-)': 'Testing and Verification Work',
            'Date of Commencement': '2024-03-01',
            'Date of Completion': '2024-12-31',
            'Date of actual completion of work :': '2024-12-20',
            'WORK ORDER AMOUNT RS.': 750000
        },
        'totals': {
            'work_order_amount': 750000,
            'last_bill_amount': 300000,
            'net_payable': 400000,
            'extra_items_sum': 25000,
            'tender_premium': 5000
        }
    }
    
    # Check if template exists
    workbook_path = "ATTACHED_ASSETS/even BILL NOTE SHEET.xlsm"
    if not os.path.exists(workbook_path):
        print(f"❌ Template workbook not found: {workbook_path}")
        return False
    
    print(f"✅ Template workbook found: {workbook_path}")
    
    # Check current sheets
    wb = openpyxl.load_workbook(workbook_path, keep_vba=True)
    print(f"📋 Current sheets in workbook: {wb.sheetnames}")
    wb.close()
    
    # Test the functionality
    print("\n🔧 Testing scrutiny sheet generation...")
    result = create_scrutiny_sheet_for_bill(
        workbook_path=workbook_path,
        processed_data=test_data,
        bill_type="running"
    )
    
    if result['success']:
        print("✅ Scrutiny sheet generation successful!")
        print(f"   Sheet Name: {result['sheet_name']}")
        print(f"   Macro Executed: {result['macro_executed']}")
        print(f"   PDF Exported: {result['pdf_exported']}")
        if result['pdf_path']:
            print(f"   PDF Path: {result['pdf_path']}")
        
        # Verify the sheet was created and data populated
        try:
            wb = openpyxl.load_workbook(workbook_path, keep_vba=True)
            if result['sheet_name'] in wb.sheetnames:
                print(f"✅ Sheet '{result['sheet_name']}' found in workbook")
                ws = wb[result['sheet_name']]
                
                # Check some key cells
                c3_value = ws['C3'].value
                c8_value = ws['C8'].value
                c9_value = ws['C9'].value
                
                print(f"   C3 (Agreement No): {c3_value}")
                print(f"   C8 (Work Name): {c8_value}")
                print(f"   C9 (Contractor): {c9_value}")
                
                # Verify data was populated
                if c3_value and c8_value and c9_value:
                    print("✅ Key data cells populated successfully")
                else:
                    print("⚠️  Some data cells may not be populated")
            else:
                print(f"❌ Sheet '{result['sheet_name']}' not found in workbook")
            wb.close()
        except Exception as e:
            print(f"⚠️  Could not verify sheet content: {e}")
        
        return True
    else:
        print(f"❌ Scrutiny sheet generation failed: {result['error']}")
        return False

if __name__ == "__main__":
    try:
        success = verify_implementation()
        if success:
            print("\n🎉 Implementation verification completed successfully!")
        else:
            print("\n💥 Implementation verification failed!")
    except Exception as e:
        print(f"\n💥 Verification failed with error: {e}")
        import traceback
        traceback.print_exc()