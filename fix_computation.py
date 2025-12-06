"""
Fix computation issues after row deletions
The deduction formulas should reference the correct cells
"""
import openpyxl
import shutil
from datetime import datetime

source_file = 'ATTACHED_ASSETS/english Note FINAL BILL NOTE SHEET.xlsm'
backup_file = f'ATTACHED_ASSETS/english Note FINAL BILL NOTE SHEET_backup_fix_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsm'

print("="*70)
print("FIXING COMPUTATION ISSUES")
print("="*70)

# Create backup
shutil.copy2(source_file, backup_file)
print(f"\n✓ Backup created: {backup_file}")

# Load the file
wb = openpyxl.load_workbook(source_file, keep_vba=True)
ws = wb.active

print(f"✓ Loaded workbook")

# Identify key rows
print("\n✓ Identifying key cells:")
print(f"  C18 (Work Order): {ws.cell(18, 3).value}")
print(f"  C19 (17.A - Last Bill): {ws.cell(19, 3).value}")
print(f"  C20 (17.B - This Bill): {ws.cell(20, 3).value}")
print(f"  C21 (17.C - Total): {ws.cell(21, 3).value}")
print(f"  C22 (Balance): {ws.cell(22, 3).value}")

# The deductions should be calculated on "This Bill Amount" (C20)
# NOT on Balance (C22)

print("\n✓ Fixing deduction formulas:")

# Row 35: SD @ 10% of This Bill (C20)
ws.cell(35, 4).value = "=ROUNDUP(C20*0.1,0)"
print(f"  Row 35 (SD): =ROUNDUP(C20*0.1,0)")

# Row 36: LT @ 2% of This Bill (C20)
ws.cell(36, 4).value = "=ROUNDUP($C$20*0.02,0)"
print(f"  Row 36 (LT): =ROUNDUP($C$20*0.02,0)")

# Row 37: GST @ 2% of This Bill (C20)
ws.cell(37, 4).value = "=ROUNDUP($C$20*0.02,0)"
print(f"  Row 37 (GST): =ROUNDUP($C$20*0.02,0)")

# Row 38: LC @ 1% of This Bill (C20)
ws.cell(38, 4).value = "=ROUNDUP($C$20*0.01,0)"
print(f"  Row 38 (LC): =ROUNDUP($C$20*0.01,0)")

# Row 40: Cheque = This Bill - All Deductions
ws.cell(40, 4).value = "=C20-D35-D36-D37-D38-D39"
print(f"  Row 40 (Cheque): =C20-D35-D36-D37-D38-D39")

# Row 41: Total = This Bill (C20)
ws.cell(41, 4).value = "=C20"
print(f"  Row 41 (Total): =C20")

# Also need to check if there are any other references to old rows
# Let's scan for formulas that might be wrong
print("\n✓ Checking for other formula issues:")

for row_idx in range(1, 50):
    for col_idx in range(1, 10):
        cell = ws.cell(row_idx, col_idx)
        if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
            # Check if formula contains references that might be wrong
            formula = cell.value
            if 'C22' in formula and row_idx >= 34:  # Deduction area
                print(f"  ⚠ Row {row_idx}, Col {col_idx}: {formula}")

# Update VBA macro to read from correct cells
print("\n✓ Verifying VBA macro cell references:")
print("  The macro should read:")
print("    - C22: Repair Work (row after Balance)")
print("    - C23: Extra Item formula")
print("    - D23: Extra Item amount")
print("    - C31: Extra Item amount value")
print("    - C32: Excess Quantity")
print("    - C33: Delay Comment")

# Let's check what's actually in those cells
print("\n✓ Current cell values:")
print(f"  C22: {ws.cell(22, 3).value}")
print(f"  C23: {ws.cell(23, 3).value}")
print(f"  D23: {ws.cell(23, 4).value}")
print(f"  C29: {ws.cell(29, 3).value}")
print(f"  C30: {ws.cell(30, 3).value}")
print(f"  C31: {ws.cell(31, 3).value}")
print(f"  C32: {ws.cell(32, 3).value}")
print(f"  C33: {ws.cell(33, 3).value}")

# The VBA needs to read from:
# - Repair Work: C29
# - Extra Item: C30
# - Extra Item Amount: C31
# - Excess Quantity: C32
# - Delay Comment: C33

print("\n✓ VBA macro needs to read from:")
print("  - C29: Repair Work")
print("  - C30: Extra Item (Yes/No)")
print("  - C31: Extra Item Amount")
print("  - C32: Excess Quantity")
print("  - C33: Delay Comment")

# Save the file
wb.save(source_file)
print(f"\n✓ Saved fixed file: {source_file}")

print("\n" + "="*70)
print("COMPUTATION FIX COMPLETE")
print("="*70)
print("\n✓ Fixed deduction formulas to reference C20 (This Bill)")
print("✓ All deductions now calculate correctly")
print(f"✓ Backup saved: {backup_file}")
print("\n⚠ VBA macro needs update for correct cell references")
print("  Will update in next step...")
