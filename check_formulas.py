import pandas as pd
from openpyxl import load_workbook
import os

# Check for formulas in the newly added sheet
file_path = "TEST_INPUT_FILES/0511-N-extra.xlsx"

if os.path.exists(file_path):
    # Load workbook
    wb = load_workbook(file_path, keep_vba=True)
    
    # Check the newly added sheet
    sheet_name = "Shree 48"
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"Checking for formulas in sheet '{sheet_name}':")
        
        # Check the cells we tried to update
        cells_to_check = ['A1', 'A2', 'A3', 'A4', 'A5', 'A6', 'A7', 'A8']
        for cell in cells_to_check:
            cell_obj = ws[cell]
            if hasattr(cell_obj, 'formula') and cell_obj.formula:
                print(f"Cell {cell} contains formula: {cell_obj.formula}")
            elif cell_obj.data_type == 'f':  # formula
                print(f"Cell {cell} contains formula (data_type=f): {cell_obj.value}")
            else:
                print(f"Cell {cell}: Value='{cell_obj.value}', Data_type='{cell_obj.data_type}'")
                
        # Check if there are any formulas in the sheet
        print("\nAll formulas in the sheet:")
        for row in ws.iter_rows():
            for cell in row:
                if hasattr(cell, 'formula') and cell.formula:
                    print(f"Cell {cell.coordinate} formula: {cell.formula}")
                elif cell.data_type == 'f':
                    print(f"Cell {cell.coordinate} formula (data_type=f): {cell.value}")
    else:
        print(f"Sheet '{sheet_name}' not found")