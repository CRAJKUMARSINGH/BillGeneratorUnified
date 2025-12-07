import pandas as pd
from core.processors.excel_processor import ExcelProcessor
from core.processors.batch_processor import add_bill_summary_sheet
import shutil
from pathlib import Path

# Process a test file
processor = ExcelProcessor()
with open('TEST_INPUT_FILES/0511-N-extra.xlsx', 'rb') as f:
    processed_data = processor.process_excel(f)

# Create a copy of the original file to test with
test_file = 'test_output/0511-N-extra_test.xlsx'
Path('test_output').mkdir(exist_ok=True)
shutil.copy2('TEST_INPUT_FILES/0511-N-extra.xlsx', test_file)

print("Processed data title keys:")
for key in processed_data['title_data'].keys():
    print(f"  '{key}'")

print("\nAdding summary sheet...")
result = add_bill_summary_sheet(test_file, processed_data)
print(f"Result: {result}")

print("\nChecking if sheet was added:")
try:
    from openpyxl import load_workbook
    wb = load_workbook(test_file)
    print(f"Sheet names: {wb.sheetnames}")
    
    # Check if our new sheet exists
    from core.processors.batch_processor import generate_sheet_name
    expected_sheet_name = generate_sheet_name(processed_data['title_data'])
    print(f"Expected sheet name: {expected_sheet_name}")
    
    if expected_sheet_name in wb.sheetnames:
        print(f"✓ Sheet '{expected_sheet_name}' found!")
        ws = wb[expected_sheet_name]
        print(f"Sheet content:")
        for row in ws.iter_rows(values_only=True):
            print(f"  {row}")
    else:
        print(f"✗ Sheet '{expected_sheet_name}' not found")
        
except Exception as e:
    print(f"Error checking workbook: {e}")