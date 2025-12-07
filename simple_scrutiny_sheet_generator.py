#!/usr/bin/env python3
"""
Simple Scrutiny Sheet Generator for Bill Processing
===============================================

This script creates scrutiny sheets for processed bills using openpyxl for most operations
and only uses COM for macro execution when available.
"""

import os
import sys
from pathlib import Path
from datetime import datetime
import re
import shutil

# Import libraries with error handling
try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    openpyxl = None
    OPENPYXL_AVAILABLE = False
    print("ERROR: openpyxl not available. Install with: pip install openpyxl")

try:
    import win32com.client
    WIN32COM_AVAILABLE = True
except ImportError:
    win32com = None
    WIN32COM_AVAILABLE = False
    print("WARNING: win32com not available. Cannot run macros. Install with: pip install pywin32")


def get_first_n_words(text, n=5):
    """
    Extract first n words from text
    
    Args:
        text (str): Input text
        n (int): Number of words to extract
        
    Returns:
        str: First n words joined by spaces
    """
    if not text or not isinstance(text, str):
        return ""
    
    # Clean and split text
    words = re.findall(r'\b\w+\b', text.strip())
    return " ".join(words[:n]) if words else ""


def sanitize_sheet_name(name):
    """
    Sanitize sheet name to comply with Excel restrictions
    
    Args:
        name (str): Sheet name to sanitize
        
    Returns:
        str: Sanitized sheet name (max 31 chars, no invalid chars)
    """
    if not name:
        return "ScrutinySheet"
    
    # Remove invalid characters for Excel sheet names
    invalid_chars = [':', '\\', '/', '?', '*', '[', ']']
    for char in invalid_chars:
        name = name.replace(char, '_')
    
    # Remove any remaining control characters
    name = ''.join(c for c in name if ord(c) >= 32)
    
    # Truncate to 31 characters (Excel limit)
    if len(name) > 31:
        name = name[:31].rstrip()
    
    return name


def generate_sheet_name(title_data):
    """
    Generate sheet name: first 5 words of contractor + "_" + agreement number
    
    Args:
        title_data (dict): Dictionary containing title sheet data
        
    Returns:
        str: Generated sheet name or None if data insufficient
    """
    # Look for contractor name variations
    contractor_keys = [
        "Name of Contractor or supplier :",
        "Name of Contractor",
        "Contractor Name",
        "Contractor",
        "Supplier Name"
    ]
    
    # Look for agreement number variations
    agreement_keys = [
        "Agreement No.",
        "Agreement No",
        "Agreement Number",
        "Agreement",
        "Work Order No",
        "Work Order Number"
    ]
    
    contractor_name = ""
    agreement_number = ""
    
    # Find contractor name
    for key in contractor_keys:
        if key in title_data and title_data[key]:
            contractor_name = str(title_data[key]).strip()
            break
    
    # Find agreement number
    for key in agreement_keys:
        if key in title_data and title_data[key]:
            agreement_number = str(title_data[key]).strip()
            break
    
    if not contractor_name or not agreement_number:
        print(f"Warning: Missing contractor ({bool(contractor_name)}) or agreement ({bool(agreement_number)}) data")
        return None
    
    # Get first 5 words of contractor name
    contractor_prefix = get_first_n_words(contractor_name, 5)
    
    # Combine with agreement number
    sheet_name = f"{contractor_prefix}_{agreement_number}"
    
    # Sanitize for Excel
    sheet_name = sanitize_sheet_name(sheet_name)
    
    return sheet_name


def populate_scrutiny_sheet_data(ws, processed_data, bill_type="running"):
    """
    Populate scrutiny sheet cells with data from processed bill
    
    Required cells:
    - C3: Agreement Number
    - C8: Name of Work
    - C9: Name of Contractor
    - C12: Date of Commencement
    - C13: Date of Completion
    - C14: "WIP" for running bills, else Actual Completion Date
    - C17: Work Order Amount
    - C18: Last Bill Amount (0 for first/final bills)
    - C19: This Bill Amount
    - C29: Sum of Extra Items
    
    Args:
        ws: Excel worksheet object
        processed_data (dict): Processed bill data
        bill_type (str): "running", "first", or "final"
    """
    title_data = processed_data.get('title_data', {})
    totals = processed_data.get('totals', {})
    
    def get_value_from_variations(data_dict, keys, default=""):
        """Helper to get value with multiple key variations"""
        for key in keys:
            if key in data_dict and data_dict[key]:
                return data_dict[key]
        return default
    
    # C3: Agreement Number
    agreement_number = get_value_from_variations(
        title_data,
        ["Agreement No.", "Agreement No", "Agreement Number", "Work Order No", "Work Order Number"],
        ""
    )
    if agreement_number:
        ws['C3'] = agreement_number
    
    # C8: Name of Work
    work_name = get_value_from_variations(
        title_data,
        ["Name of Work ;-)", "Name of Work", "Project Name", "Work Description", "Name of Work ;-", "Work Name"],
        ""
    )
    if work_name:
        ws['C8'] = work_name
    
    # C9: Name of Contractor
    contractor_name = get_value_from_variations(
        title_data,
        ["Name of Contractor or supplier :", "Name of Contractor", "Contractor Name", "Supplier Name", "Contractor"],
        ""
    )
    if contractor_name:
        ws['C9'] = contractor_name
    
    # C12: Date of Commencement
    commencement_date = get_value_from_variations(
        title_data,
        ["Date of Commencement", "Start Date", "Commencement Date", "St. date of Start :", "Start Date :"],
        ""
    )
    if commencement_date:
        ws['C12'] = commencement_date
    
    # C13: Date of Completion
    completion_date = get_value_from_variations(
        title_data,
        ["Date of Completion", "Schedule Completion", "Completion Date", "St. date of completion :", "Completion Date :"],
        ""
    )
    if completion_date:
        ws['C13'] = completion_date
    
    # C14: WIP for running bills, else actual completion date
    if bill_type.lower() == "running":
        ws['C14'] = "WIP"  # WORK IN PROGRESS
    else:
        actual_completion = get_value_from_variations(
            title_data,
            ["Date of actual completion of work :", "Actual Date of Completion", "Actual Completion", "Actual Completion Date"],
            ""
        )
        ws['C14'] = actual_completion if actual_completion else completion_date
    
    # C17: Work Order Amount
    work_order_amount = totals.get('work_order_amount', 0)
    if not work_order_amount:
        work_order_amount_str = get_value_from_variations(
            title_data,
            ["WORK ORDER AMOUNT RS.", "Work Order Amount", "Work Order Value", "Work Order Amount Rs."],
            "0"
        )
        try:
            work_order_amount = float(work_order_amount_str) if work_order_amount_str else 0
        except (ValueError, TypeError):
            work_order_amount = 0
    ws['C17'] = work_order_amount
    
    # C18: Last Bill Amount (0 for first/final bills)
    if bill_type.lower() in ["first", "final"]:
        ws['C18'] = 0
    else:
        last_bill_amount = totals.get('last_bill_amount', 0)
        # If not available in totals, try to get from title data
        if not last_bill_amount:
            last_bill_amount_str = get_value_from_variations(
                title_data,
                ["Amount Paid Vide Last Bill", "Last Bill Amount", "Previous Bill Amount"],
                "0"
            )
            try:
                last_bill_amount = float(last_bill_amount_str) if last_bill_amount_str else 0
            except (ValueError, TypeError):
                last_bill_amount = 0
        ws['C18'] = last_bill_amount if last_bill_amount else 0
    
    # C19: This Bill Amount
    this_bill_amount = totals.get('net_payable', totals.get('grand_total', 0))
    if not this_bill_amount:
        this_bill_amount = totals.get('this_bill_amount', 0)
    if not this_bill_amount:
        this_bill_amount_str = get_value_from_variations(
            title_data,
            ["This Bill Amount", "Current Bill Amount"],
            "0"
        )
        try:
            this_bill_amount = float(this_bill_amount_str) if this_bill_amount_str else 0
        except (ValueError, TypeError):
            this_bill_amount = 0
    ws['C19'] = this_bill_amount
    
    # C29: Sum of Extra Items including Tender Premium
    extra_items_sum = totals.get('extra_items_sum', 0)
    tender_premium = totals.get('tender_premium', 0)
    
    # If not in totals, try to get from title data
    if not extra_items_sum:
        extra_items_sum_str = get_value_from_variations(
            title_data,
            ["Extra Items Amount", "Sum of Extra Items"],
            "0"
        )
        try:
            extra_items_sum = float(extra_items_sum_str) if extra_items_sum_str else 0
        except (ValueError, TypeError):
            extra_items_sum = 0
    
    if not tender_premium:
        tender_premium_str = get_value_from_variations(
            title_data,
            ["Tender Premium", "Premium Amount"],
            "0"
        )
        try:
            tender_premium = float(tender_premium_str) if tender_premium_str else 0
        except (ValueError, TypeError):
            tender_premium = 0
    
    total_extra = float(extra_items_sum) + float(tender_premium)
    ws['C29'] = total_extra


def copy_template_sheet_simple(template_path, output_path, new_sheet_name):
    """
    Copy template workbook and rename the first sheet using openpyxl
    
    Args:
        template_path (str): Path to template workbook
        output_path (str): Path for output workbook
        new_sheet_name (str): New name for the first sheet
        
    Returns:
        bool: True if successful, False otherwise
    """
    if not OPENPYXL_AVAILABLE or openpyxl is None:
        print("ERROR: openpyxl not available.")
        return False
    
    try:
        # Copy the template file
        shutil.copy2(template_path, output_path)
        
        # Load the copied workbook
        wb = openpyxl.load_workbook(output_path, keep_vba=True)
        
        # Rename the first sheet
        if wb.sheetnames:
            old_sheet_name = wb.sheetnames[0]
            ws = wb[old_sheet_name]
            ws.title = new_sheet_name
            print(f"Renamed sheet '{old_sheet_name}' to '{new_sheet_name}'")
        
        # Save the workbook
        wb.save(output_path)
        wb.close()
        
        print(f"Created workbook: {output_path}")
        return True
        
    except Exception as e:
        print(f"ERROR copying template: {e}")
        return False


def run_macro_if_available(workbook_path, sheet_name):
    """
    Run macro if COM is available
    
    Args:
        workbook_path (str): Path to Excel workbook
        sheet_name (str): Name of sheet
        
    Returns:
        bool: True if macro executed, False otherwise
    """
    if not WIN32COM_AVAILABLE or win32com is None:
        print("WARNING: win32com not available. Skipping macro execution.")
        return False
    
    excel = None
    wb = None
    try:
        # Create Excel application
        excel = win32com.client.Dispatch("Excel.Application")
        if excel is None:
            print("ERROR: Could not create Excel application")
            return False
            
        excel.Visible = False
        excel.DisplayAlerts = False
        
        # Open workbook
        wb = excel.Workbooks.Open(os.path.abspath(workbook_path))
        
        # Activate the sheet
        ws = wb.Worksheets(sheet_name)
        ws.Activate()
        
        # Try to run the macro
        macro_executed = False
        try:
            excel.Application.Run("GenerateBillNotes")
            macro_executed = True
            print("Executed macro: GenerateBillNotes")
        except:
            try:
                excel.Application.Run(f"{wb.Name}!GenerateBillNotes")
                macro_executed = True
                print(f"Executed macro: {wb.Name}!GenerateBillNotes")
            except:
                pass
        
        # Save and close
        wb.Save()
        wb.Close(SaveChanges=True)
        excel.Quit()
        
        if not macro_executed:
            print("WARNING: Could not execute macro automatically.")
            return False
            
        return True
        
    except Exception as e:
        print(f"ERROR running macro: {e}")
        try:
            if wb:
                wb.Close(SaveChanges=False)
        except:
            pass
        try:
            if excel:
                excel.Quit()
        except:
            pass
        return False


def export_to_pdf_if_available(workbook_path, sheet_name, output_pdf_path):
    """
    Export to PDF if COM is available
    
    Args:
        workbook_path (str): Path to Excel workbook
        sheet_name (str): Name of sheet to export
        output_pdf_path (str): Path for output PDF file
        
    Returns:
        bool: True if successful, False otherwise
    """
    if not WIN32COM_AVAILABLE or win32com is None:
        print("WARNING: win32com not available. Skipping PDF export.")
        return False
    
    excel = None
    wb = None
    try:
        # Create Excel application
        excel = win32com.client.Dispatch("Excel.Application")
        if excel is None:
            print("ERROR: Could not create Excel application")
            return False
            
        excel.Visible = False
        excel.DisplayAlerts = False
        
        # Open workbook
        wb = excel.Workbooks.Open(os.path.abspath(workbook_path))
        
        # Activate the sheet
        ws = wb.Worksheets(sheet_name)
        ws.Activate()
        
        # Export to PDF
        ws.ExportAsFixedFormat(
            Type=0,  # xlTypePDF
            Filename=os.path.abspath(output_pdf_path),
            Quality=0,  # xlQualityStandard
            IncludeDocProperties=True,
            IgnorePrintAreas=False,
            OpenAfterPublish=False
        )
        
        # Close
        wb.Close(SaveChanges=False)
        excel.Quit()
        
        print(f"Exported to PDF: {output_pdf_path}")
        return True
        
    except Exception as e:
        print(f"ERROR exporting to PDF: {e}")
        try:
            if wb:
                wb.Close(SaveChanges=False)
        except:
            pass
        try:
            if excel:
                excel.Quit()
        except:
            pass
        return False


def create_scrutiny_sheet_simple(
    template_path,
    output_path,
    processed_data,
    bill_type="running",
    output_pdf_path=None
):
    """
    Simple function to create scrutiny sheet for a processed bill
    
    Args:
        template_path (str): Path to template workbook (.xlsm)
        output_path (str): Path for output workbook
        processed_data (dict): Processed bill data
        bill_type (str): "running", "first", or "final"
        output_pdf_path (str): Path for PDF output (optional)
        
    Returns:
        dict: Result with success status and details
    """
    if not OPENPYXL_AVAILABLE or openpyxl is None:
        return {
            'success': False,
            'error': "Required library 'openpyxl' not available. Install with: pip install openpyxl"
        }
    
    try:
        # Generate sheet name
        title_data = processed_data.get('title_data', {})
        sheet_name = generate_sheet_name(title_data)
        
        if not sheet_name:
            return {
                'success': False,
                'error': "Could not generate sheet name - missing contractor or agreement data"
            }
        
        print(f"Generated sheet name: {sheet_name}")
        
        # Copy template and rename sheet
        if not copy_template_sheet_simple(template_path, output_path, sheet_name):
            return {
                'success': False,
                'error': "Failed to copy template sheet"
            }
        
        # Load workbook and populate data
        try:
            wb = openpyxl.load_workbook(output_path, keep_vba=True)
            ws = wb[sheet_name]
            
            # Populate data
            print("Populating sheet data...")
            populate_scrutiny_sheet_data(ws, processed_data, bill_type)
            
            # Save workbook
            wb.save(output_path)
            wb.close()
            print("Data populated successfully")
            
        except Exception as e:
            print(f"ERROR populating data: {e}")
            return {
                'success': False,
                'error': f"Failed to populate sheet data: {e}"
            }
        
        # Run macro if available
        print("Attempting to run macro...")
        macro_success = run_macro_if_available(output_path, sheet_name)
        
        # Export to PDF if requested and available
        pdf_success = False
        if output_pdf_path:
            print("Attempting to export to PDF...")
            pdf_success = export_to_pdf_if_available(output_path, sheet_name, output_pdf_path)
        
        return {
            'success': True,
            'sheet_name': sheet_name,
            'output_path': str(output_path),
            'pdf_path': str(output_pdf_path) if pdf_success else None,
            'macro_executed': macro_success,
            'pdf_exported': pdf_success
        }
        
    except Exception as e:
        return {
            'success': False,
            'error': str(e)
        }


# Example usage
if __name__ == "__main__":
    print("Simple Scrutiny Sheet Generator")
    print("===============================")
    
    # Example data structure (this would come from your bill processing)
    example_processed_data = {
        'title_data': {
            'Name of Contractor or supplier :': 'M/s. Shree Krishna Builders Jaipur',
            'Agreement No.': '48/2024-25',
            'Name of Work ;-)': 'Plumbing Installation and MTC work at Govt. Nehru hostel Mansarovar, Sanganer, Jaipur',
            'Date of Commencement': '2024-01-15',
            'Date of Completion': '2024-12-15',
            'Date of actual completion of work :': '2024-12-10',
            'WORK ORDER AMOUNT RS.': 500000
        },
        'totals': {
            'work_order_amount': 500000,
            'last_bill_amount': 200000,
            'net_payable': 250000,
            'extra_items_sum': 15000,
            'tender_premium': 0
        }
    }
    
    # Example usage
    template_path = "ATTACHED_ASSETS/even BILL NOTE SHEET.xlsm"
    output_path = "test_output/simple_scrutiny_sheet.xlsm"
    pdf_path = "test_output/MACRO scrutiny SHEET IN PDF.pdf"
    
    # Create output directory
    Path("test_output").mkdir(exist_ok=True)
    
    if os.path.exists(template_path):
        result = create_scrutiny_sheet_simple(
            template_path=template_path,
            output_path=output_path,
            processed_data=example_processed_data,
            bill_type="running",
            output_pdf_path=pdf_path
        )
        
        if result['success']:
            print(f"\n✅ SUCCESS!")
            print(f"   Sheet Name: {result['sheet_name']}")
            print(f"   Output File: {result['output_path']}")
            print(f"   PDF Exported: {result['pdf_exported']}")
            if result['pdf_path']:
                print(f"   PDF Path: {result['pdf_path']}")
            print(f"   Macro Executed: {result['macro_executed']}")
        else:
            print(f"\n❌ FAILED: {result['error']}")
    else:
        print(f"\n❌ Template not found: {template_path}")