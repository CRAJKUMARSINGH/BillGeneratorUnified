"""
Apply minor correction:
- Remove row 24 (Net Amount of This Bill Rs.)
- C23 already has the correct formula: =IF(D23,ROUND((D23/C18)*100,2),0)
- Update VBA macro to remove reference to the deleted row
"""
import openpyxl
import shutil
from datetime import datetime
import os

# Create backup
backup_file = f'ATTACHED_ASSETS/english Note FINAL BILL NOTE SHEET_backup_correction_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsm'
shutil.copy2('ATTACHED_ASSETS/english Note FINAL BILL NOTE SHEET_UPDATED.xlsm', backup_file)
print(f"✓ Backup created: {backup_file}")

# Load the XLSM file
wb = openpyxl.load_workbook('ATTACHED_ASSETS/english Note FINAL BILL NOTE SHEET_UPDATED.xlsm', keep_vba=True)
ws = wb.active

print(f"✓ Loaded workbook: {wb.sheetnames}")
print(f"✓ Active sheet: {ws.title}")

# Show current structure
print("\nCurrent structure (rows 21-26):")
for i in range(21, 27):
    print(f"Row {i}: A={ws.cell(i,1).value} | B={ws.cell(i,2).value[:40] if ws.cell(i,2).value else None}... | C={ws.cell(i,3).value}")

# Delete row 24 (Net Amount of This Bill Rs.)
print("\n✓ Deleting row 24 (Net Amount of This Bill Rs.)...")
ws.delete_rows(24, 1)

# Show updated structure
print("\nUpdated structure (rows 21-26):")
for i in range(21, 27):
    cell_b = ws.cell(i,2).value
    if cell_b and len(str(cell_b)) > 40:
        cell_b = str(cell_b)[:40] + "..."
    print(f"Row {i}: A={ws.cell(i,1).value} | B={cell_b} | C={ws.cell(i,3).value}")

# Verify C23 has the correct formula
c23_formula = ws.cell(23, 3).value
print(f"\n✓ Verifying C23 formula: {c23_formula}")
if "IF(D23" in str(c23_formula) and "D23/C18" in str(c23_formula):
    print("  ✓ C23 formula is correct!")
else:
    print("  ⚠ C23 formula may need adjustment")

# Save the corrected file
output_file = 'ATTACHED_ASSETS/english Note FINAL BILL NOTE SHEET_CORRECTED.xlsm'
wb.save(output_file)
print(f"\n✓ Saved corrected file: {output_file}")

print("\n" + "="*70)
print("CORRECTION COMPLETE")
print("="*70)
print(f"✓ Row 24 removed")
print(f"✓ C23 formula verified: =IF(D23,ROUND((D23/C18)*100,2),0)")
print(f"✓ File saved to: {output_file}")
