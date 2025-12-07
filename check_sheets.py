import pandas as pd
from openpyxl import load_workbook
import os

# Check the sheets in one of the Excel files
file_path = "TEST_INPUT_FILES/0511-N-extra.xlsx"

if os.path.exists(file_path):
    # Load workbook
    wb = load_workbook(file_path, keep_vba=True)
    print("Sheets in the workbook:")
    for sheet_name in wb.sheetnames:
        print(f"  - {sheet_name}")
    
    # Check if it's a macro-enabled file
    print(f"\nFile type: {'Macro-enabled' if wb.vba_archive else 'Regular'}")
    
    # Show some info about each sheet
    print("\nSheet details:")
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"  {sheet_name}: {ws.max_row} rows, {ws.max_column} columns")
else:
    print(f"File {file_path} not found")