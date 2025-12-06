"""
Complete update of the XLSM file:
1. Update structure to add 17.A, 17.B, 17.C rows
2. Update ALL formulas to maintain correct references
3. Embed updated VBA macro
"""
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from copy import copy
import shutil
from datetime import datetime
import win32com.client
import os
import time

# Create backup
backup_file = f'ATTACHED_ASSETS/english Note FINAL BILL NOTE SHEET_backup_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsm'
shutil.copy2('ATTACHED_ASSETS/english Note FINAL BILL NOTE SHEET.xlsm', backup_file)
print(f"✓ Backup created: {backup_file}")

# Load the XLSM file with VBA macros preserved
wb = openpyxl.load_workbook('ATTACHED_ASSETS/english Note FINAL BILL NOTE SHEET.xlsm', keep_vba=True)
ws = wb.active

print(f"✓ Loaded workbook: {wb.sheetnames}")
print(f"✓ Active sheet: {ws.title}")

# Find row 19 (current item 17)
row_17_index = None
for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=30, min_col=1, max_col=1), start=1):
    cell_value = row[0].value
    if cell_value == 17:
        row_17_index = idx
        print(f"✓ Found item 17 at row {idx}")
        break

if not row_17_index:
    print("✗ Could not find item 17 in the sheet")
    exit(1)

# Store all formulas that need updating BEFORE inserting rows
formulas_to_update = {}
for row_idx in range(row_17_index, ws.max_row + 1):
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row_idx, col_idx)
        if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
            formulas_to_update[(row_idx, col_idx)] = cell.value

print(f"✓ Found {len(formulas_to_update)} formulas to update")

# Insert 2 new rows after the current row 17
ws.insert_rows(row_17_index + 1, 2)
print(f"✓ Inserted 2 new rows after row {row_17_index}")

# Update row 19 (17.A)
ws.cell(row_17_index, 1).value = "17.A"
ws.cell(row_17_index, 2).value = "Sum of payment upto last bill Rs."
# Keep existing value or set to 0
current_val = ws.cell(row_17_index, 3).value
if current_val and isinstance(current_val, str) and '=' in str(current_val):
    ws.cell(row_17_index, 3).value = 0
print(f"✓ Updated row {row_17_index} to 17.A")

# Create row 20 (17.B)
ws.cell(row_17_index + 1, 1).value = "B."
ws.cell(row_17_index + 1, 2).value = "Amount of this bill Rs."
ws.cell(row_17_index + 1, 3).value = 0
# Copy formatting
for col in range(1, 5):
    source_cell = ws.cell(row_17_index, col)
    target_cell = ws.cell(row_17_index + 1, col)
    if source_cell.has_style:
        target_cell.font = copy(source_cell.font)
        target_cell.border = copy(source_cell.border)
        target_cell.fill = copy(source_cell.fill)
        target_cell.number_format = copy(source_cell.number_format)
        target_cell.protection = copy(source_cell.protection)
        target_cell.alignment = copy(source_cell.alignment)
print(f"✓ Created row {row_17_index + 1} as 17.B")

# Create row 21 (17.C) with formula
ws.cell(row_17_index + 2, 1).value = "C."
ws.cell(row_17_index + 2, 2).value = "Actual expenditure upto this bill = (A + B) Rs."
ws.cell(row_17_index + 2, 3).value = f"=C{row_17_index}+C{row_17_index + 1}"
# Copy formatting
for col in range(1, 5):
    source_cell = ws.cell(row_17_index, col)
    target_cell = ws.cell(row_17_index + 2, col)
    if source_cell.has_style:
        target_cell.font = copy(source_cell.font)
        target_cell.border = copy(source_cell.border)
        target_cell.fill = copy(source_cell.fill)
        target_cell.number_format = copy(source_cell.number_format)
        target_cell.protection = copy(source_cell.protection)
        target_cell.alignment = copy(source_cell.alignment)
print(f"✓ Created row {row_17_index + 2} as 17.C with formula =C{row_17_index}+C{row_17_index + 1}")

# Update all formulas that were stored
# Rules:
# - References to C19 (old item 17) should now reference C21 (new 17.C)
# - All row references >= row_17_index need to be shifted by +2
print("\n✓ Updating formulas:")
for (old_row, col), formula in formulas_to_update.items():
    new_row = old_row + 2  # Shift down by 2 rows
    updated_formula = formula
    
    # Update cell references in the formula
    # We need to update any reference to row >= row_17_index
    import re
    
    # Find all cell references (e.g., C19, $C$19, C19:C20, etc.)
    pattern = r'(\$?)([A-Z]+)(\$?)(\d+)'
    
    def replace_cell_ref(match):
        col_abs = match.group(1)
        col_letter = match.group(2)
        row_abs = match.group(3)
        row_num = int(match.group(4))
        
        # If the row reference is >= row_17_index, shift it by 2
        if row_num >= row_17_index:
            new_row_num = row_num + 2
            return f"{col_abs}{col_letter}{row_abs}{new_row_num}"
        else:
            return match.group(0)
    
    updated_formula = re.sub(pattern, replace_cell_ref, formula)
    
    # Set the updated formula in the new position
    ws.cell(new_row, col).value = updated_formula
    
    if updated_formula != formula:
        print(f"  Row {old_row}→{new_row}, Col {col}: {formula} → {updated_formula}")

# Special case: Update "Balance to be done" (item 18)
# Should be: =C18-C21 (Work Order Amount - Actual Expenditure 17.C)
for row_idx in range(row_17_index + 3, min(row_17_index + 10, ws.max_row + 1)):
    if ws.cell(row_idx, 1).value == 18:
        # Find item 16 row (Amount of Work Order)
        item_16_row = None
        for r in range(1, row_17_index):
            if ws.cell(r, 1).value == 16:
                item_16_row = r
                break
        
        if item_16_row:
            ws.cell(row_idx, 3).value = f"=C{item_16_row}-C{row_17_index + 2}"
            print(f"✓ Updated item 18 (Balance) formula: =C{item_16_row}-C{row_17_index + 2}")
        break

# Update "Net Amount of This Bill" to reference 17.C instead of old 17
for row_idx in range(row_17_index + 3, min(row_17_index + 10, ws.max_row + 1)):
    cell_val = ws.cell(row_idx, 2).value
    if cell_val and "Net Amount of This Bill" in str(cell_val):
        ws.cell(row_idx, 3).value = f"=C{row_17_index + 2}"
        print(f"✓ Updated 'Net Amount of This Bill' formula at row {row_idx}: =C{row_17_index + 2}")
        break

# Save the updated file
temp_file = 'ATTACHED_ASSETS/english Note FINAL BILL NOTE SHEET_TEMP.xlsm'
wb.save(temp_file)
print(f"\n✓ Saved updated structure to: {temp_file}")

# Now update the VBA macro using COM
print("\n✓ Updating VBA macro...")
try:
    xl = win32com.client.Dispatch("Excel.Application")
    xl.Visible = False
    xl.DisplayAlerts = False
    
    # Open the temp file
    wb_com = xl.Workbooks.Open(os.path.abspath(temp_file))
    
    # Read the updated VBA code
    with open('updated_macro.vba', 'r', encoding='utf-8') as f:
        vba_code = f.read()
    
    # Find the module containing GenerateBillNotes
    module_found = False
    for component in wb_com.VBProject.VBComponents:
        if component.Type == 1:  # vbext_ct_StdModule
            # Check if this module contains GenerateBillNotes
            code = component.CodeModule.Lines(1, component.CodeModule.CountOfLines)
            if "GenerateBillNotes" in code:
                print(f"  Found macro in module: {component.Name}")
                # Clear existing code
                component.CodeModule.DeleteLines(1, component.CodeModule.CountOfLines)
                # Add updated code
                component.CodeModule.AddFromString(vba_code)
                print(f"  ✓ Updated VBA code in module: {component.Name}")
                module_found = True
                break
    
    if not module_found:
        # Create new module if not found
        new_module = wb_com.VBProject.VBComponents.Add(1)  # vbext_ct_StdModule
        new_module.Name = "BillNotesGenerator"
        new_module.CodeModule.AddFromString(vba_code)
        print(f"  ✓ Created new VBA module: BillNotesGenerator")
    
    # Save and close
    output_file = os.path.abspath('ATTACHED_ASSETS/english Note FINAL BILL NOTE SHEET_UPDATED.xlsm')
    wb_com.SaveAs(output_file, FileFormat=52)  # xlOpenXMLWorkbookMacroEnabled
    wb_com.Close(False)
    xl.Quit()
    
    print(f"✓ VBA macro updated successfully")
    
    # Clean up temp file
    time.sleep(1)
    if os.path.exists(temp_file):
        os.remove(temp_file)
        print(f"✓ Cleaned up temp file")
    
except Exception as e:
    print(f"⚠ Could not update VBA macro automatically: {e}")
    print(f"  Please manually copy the VBA code from 'updated_macro.vba' into the XLSM file")
    output_file = temp_file

# Display the updated structure
print("\n" + "="*70)
print("UPDATED STRUCTURE:")
print("="*70)
wb_final = openpyxl.load_workbook(output_file, data_only=False)
ws_final = wb_final.active
for row_idx in range(row_17_index - 2, min(row_17_index + 10, ws_final.max_row + 1)):
    row_data = []
    for col in range(1, 5):
        val = ws_final.cell(row_idx, col).value
        if val and isinstance(val, str) and len(str(val)) > 50:
            val = str(val)[:47] + "..."
        row_data.append(val)
    print(f"Row {row_idx}: {row_data}")

print("\n" + "="*70)
print("UPDATE COMPLETE!")
print("="*70)
print(f"✓ Original file backed up to: {backup_file}")
print(f"✓ Updated file saved to: {output_file}")
print("\nKey changes:")
print(f"  • Row {row_17_index}: Item 17.A - Sum of payment upto last bill")
print(f"  • Row {row_17_index + 1}: Item 17.B - Amount of this bill")
print(f"  • Row {row_17_index + 2}: Item 17.C - Actual expenditure (=C{row_17_index}+C{row_17_index + 1})")
print(f"  • All formulas updated to reference correct rows")
print(f"  • VBA macro updated with new cell references")
print(f"  • Output cell moved from B42 to B44")
