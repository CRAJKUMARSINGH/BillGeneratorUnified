"""
Convert EVEN BILL_SCRUTINY_SHEET.xlsx to .xlsm and add VBA macro
"""
import openpyxl
import win32com.client
import os
import shutil
from datetime import datetime

source_file = 'ATTACHED_ASSETS/EVEN BILL_SCRUTINY_SHEET.xlsx'
output_file = 'ATTACHED_ASSETS/EVEN BILL_SCRUTINY_SHEET.xlsm'

print("="*70)
print("ADDING VBA MACRO TO EVEN BILL_SCRUTINY_SHEET")
print("="*70)

# First, check the structure
wb = openpyxl.load_workbook(source_file)
ws = wb.active

print(f"\n✓ Loaded file: {source_file}")
print(f"✓ Active sheet: {ws.title}")

# Show structure
print("\nCurrent structure (rows 17-26):")
for i in range(17, 27):
    cell_a = ws.cell(i, 1).value
    if cell_a and len(str(cell_a)) > 50:
        cell_a = str(cell_a)[:50] + "..."
    cell_c = ws.cell(i, 3).value
    if cell_c and isinstance(cell_c, str) and len(cell_c) > 30:
        cell_c = cell_c[:30] + "..."
    print(f"Row {i}: {cell_a} | C={cell_c}")

# Check if we need to update any cell references for the macro
# The macro expects certain cells, let's verify:
print("\n✓ Verifying key cells for macro:")
print(f"  C5 (Agreement): {ws.cell(5, 3).value}")
print(f"  C13 (Start Date): {ws.cell(13, 3).value}")
print(f"  C14 (Schedule Completion): {ws.cell(14, 3).value}")
print(f"  C15 (Actual Completion): {ws.cell(15, 3).value}")
print(f"  C18 (Work Order Amount): {ws.cell(18, 3).value}")
print(f"  C19 (17.A - Upto Last Bill): {ws.cell(19, 3).value}")
print(f"  C20 (17.B - This Bill): {ws.cell(20, 3).value}")
print(f"  C21 (17.C - Total): {ws.cell(21, 3).value}")

# Now we need to find where the other fields are
# Let's search for them
print("\n✓ Searching for other required fields...")

# We need to find: repair work, extra item, extra item amount, excess quantity, delay comment
# Let's scan the sheet
for i in range(1, 40):
    cell_val = ws.cell(i, 1).value
    if cell_val:
        cell_str = str(cell_val).lower()
        if 'repair' in cell_str or 'deposit' in cell_str:
            print(f"  Row {i}: {cell_val} | C={ws.cell(i, 3).value}")
        elif 'extra' in cell_str:
            print(f"  Row {i}: {cell_val} | C={ws.cell(i, 3).value}")
        elif 'excess' in cell_str or 'quantity' in cell_str:
            print(f"  Row {i}: {cell_val} | C={ws.cell(i, 3).value}")

wb.close()

# Now convert to XLSM and add VBA using COM
print("\n" + "="*70)
print("CONVERTING TO XLSM AND ADDING VBA MACRO")
print("="*70)

try:
    xl = win32com.client.Dispatch("Excel.Application")
    xl.Visible = False
    xl.DisplayAlerts = False
    
    # Open the xlsx file
    wb_com = xl.Workbooks.Open(os.path.abspath(source_file))
    
    # Read the VBA code
    with open('updated_macro_corrected.vba', 'r', encoding='utf-8') as f:
        vba_code = f.read()
    
    print("✓ Loaded VBA code from: updated_macro_corrected.vba")
    
    # Add a new module
    new_module = wb_com.VBProject.VBComponents.Add(1)  # 1 = vbext_ct_StdModule
    new_module.Name = "BillNotesGenerator"
    new_module.CodeModule.AddFromString(vba_code)
    print(f"✓ Created VBA module: {new_module.Name}")
    
    # Save as XLSM
    wb_com.SaveAs(os.path.abspath(output_file), FileFormat=52)  # 52 = xlOpenXMLWorkbookMacroEnabled
    print(f"✓ Saved as XLSM: {output_file}")
    
    wb_com.Close(False)
    xl.Quit()
    
    print("\n" + "="*70)
    print("✅ SUCCESS!")
    print("="*70)
    print(f"✓ Created: {output_file}")
    print(f"✓ VBA macro 'GenerateBillNotes' added")
    print("\nThe file now has:")
    print("  • Correct structure with 17.A, 17.B, 17.C")
    print("  • VBA macro for generating bill notes")
    print("  • Output will be in cell B43")
    print("\nTo use:")
    print("  1. Open the .xlsm file")
    print("  2. Enable macros")
    print("  3. Fill in the data")
    print("  4. Press Alt+F8, select 'GenerateBillNotes', Run")
    
except Exception as e:
    print(f"\n✗ Error: {e}")
    print("\nIf VBA access is not trusted, please:")
    print("  1. Open Excel")
    print("  2. File → Options → Trust Center → Trust Center Settings")
    print("  3. Macro Settings → Check 'Trust access to VBA project object model'")
    print("  4. Run this script again")
    try:
        xl.Quit()
    except:
        pass
