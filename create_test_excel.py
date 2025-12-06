import pandas as pd
from openpyxl import Workbook

def create_test_excel():
    # Create a new workbook
    wb = Workbook()
    
    # Remove the default sheet
    wb.remove(wb.active)
    
    # Create Title sheet
    title_sheet = wb.create_sheet("Title")
    title_sheet['A1'] = 'Project Name'
    title_sheet['B1'] = 'Test Project'
    title_sheet['A2'] = 'Work Order No.'
    title_sheet['B2'] = 'WO-001'
    title_sheet['A3'] = 'Date'
    title_sheet['B3'] = '2025-12-04'
    
    # Create Work Order sheet
    work_order_sheet = wb.create_sheet("Work Order")
    work_order_sheet['A1'] = 'Item'
    work_order_sheet['B1'] = 'Description'
    work_order_sheet['C1'] = 'Quantity'
    work_order_sheet['D1'] = 'Rate'
    work_order_sheet['E1'] = 'Amount'
    
    work_order_sheet['A2'] = 'Item 1'
    work_order_sheet['B2'] = 'Test Item Description'
    work_order_sheet['C2'] = 10
    work_order_sheet['D2'] = 100
    work_order_sheet['E2'] = 1000
    
    work_order_sheet['A3'] = 'Item 2'
    work_order_sheet['B3'] = 'Another Test Item'
    work_order_sheet['C3'] = 5
    work_order_sheet['D3'] = 50
    work_order_sheet['E3'] = 250
    
    # Create Bill Quantity sheet
    bill_qty_sheet = wb.create_sheet("Bill Quantity")
    bill_qty_sheet['A1'] = 'Sl.No.'
    bill_qty_sheet['B1'] = 'Description'
    bill_qty_sheet['C1'] = 'Unit'
    bill_qty_sheet['D1'] = 'Quantity'
    
    bill_qty_sheet['A2'] = 1
    bill_qty_sheet['B2'] = 'Test Work'
    bill_qty_sheet['C2'] = 'Nos'
    bill_qty_sheet['D2'] = 15
    
    # Save the workbook
    wb.save("input/test_project.xlsx")
    print("✅ Created test_project.xlsx in input folder")

if __name__ == "__main__":
    create_test_excel()