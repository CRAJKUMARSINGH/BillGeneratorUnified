"""
Fix note sheet computing - adjust cell references for proper calculation
The issue is that row 23 formula references D23 but the actual extra item amount is in C31
"""
import openpyxl
import shutil
from datetime import datetime

source_file = 'ATTACHED_ASSETS/english Note FINAL BILL NOTE SHEET.xlsm'
backup_file = f'ATTACHED_ASSETS/english Note FINAL BILL NOTE SHEET_backup_notesheet_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsm'

print("="*70)
print("FIXING NOTE SHEET COMPUTING")
print("="*70)

# Create backup
shutil.copy2(source_file, backup_file)
print(f"\n✓ Backup created: {backup_file}")

# Load the file
wb = openpyxl.load_workbook(source_file, keep_vba=True)
ws = wb.active

print(f"✓ Loaded workbook")

# Check current structure
print("\n✓ Current structure:")
print(f"  Row 23 C23: {ws.cell(23, 3).value}")
print(f"  Row 23 D23: {ws.cell(23, 4).value}")
print(f"  Row 31 C31 (Extra Item Amount): {ws.cell(31, 3).value}")

# The formula in C23 should calculate percentage based on C31 (not D23)
# C23 should show: =IF(C31,ROUND((C31/C18)*100,2),0)
# This will calculate the percentage of extra items

print("\n✓ Fixing row 23 formula:")
ws.cell(23, 3).value = "=IF(C31,ROUND((C31/C18)*100,2),0)"
print(f"  C23: =IF(C31,ROUND((C31/C18)*100,2),0)")
print(f"  This calculates: (Extra Item Amount / Work Order) × 100")

# Also check if there are any other formulas that need adjustment
print("\n✓ Checking other formulas:")

# Row 14 has date calculations - these look fine
print(f"  E14 (Days): {ws.cell(14, 5).value}")
print(f"  F14 (Half): {ws.cell(14, 6).value}")

# Check if row 23 should have a label in column B
if ws.cell(23, 2).value:
    print(f"  Row 23 label: {ws.cell(23, 2).value}")

# Save the file
wb.save(source_file)
print(f"\n✓ Saved fixed file: {source_file}")

# Verify the fix
wb_check = openpyxl.load_workbook(source_file, data_only=False)
ws_check = wb_check.active

print("\n" + "="*70)
print("VERIFICATION:")
print("="*70)
print(f"\n✓ C23 formula: {ws_check.cell(23, 3).value}")
print(f"✓ C31 value: {ws_check.cell(31, 3).value}")

# Test calculation
work_order = ws_check.cell(18, 3).value
extra_amount = ws_check.cell(31, 3).value
if work_order and extra_amount:
    expected_percentage = (extra_amount / work_order) * 100
    print(f"\n✓ Expected calculation:")
    print(f"  Work Order (C18): {work_order}")
    print(f"  Extra Amount (C31): {extra_amount}")
    print(f"  Percentage: {expected_percentage:.2f}%")

print("\n" + "="*70)
print("NOTE SHEET COMPUTING FIX COMPLETE")
print("="*70)
print(f"\n✓ Formula updated: C23 now references C31 (Extra Item Amount)")
print(f"✓ Backup saved: {backup_file}")
print("\n📝 The note sheet will now compute correctly!")
