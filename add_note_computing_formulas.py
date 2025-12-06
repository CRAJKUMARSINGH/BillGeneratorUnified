"""
Add note computing formulas at A40 area
These formulas will calculate:
- Percentage of work done
- Delay days
- Extra item percentage
- Time allowed
"""
import openpyxl
import shutil
from datetime import datetime

source_file = 'ATTACHED_ASSETS/english Note FINAL BILL NOTE SHEET.xlsm'
backup_file = f'ATTACHED_ASSETS/english Note FINAL BILL NOTE SHEET_backup_notecompute_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsm'

print("="*70)
print("ADDING NOTE COMPUTING FORMULAS")
print("="*70)

# Create backup
shutil.copy2(source_file, backup_file)
print(f"\n✓ Backup created: {backup_file}")

# Load the file
wb = openpyxl.load_workbook(source_file, keep_vba=True)
ws = wb.active

print(f"✓ Loaded workbook")

# Add computing formulas in a hidden area (rows 34-39, columns E-F)
# These will be used by the VBA macro or for reference

print("\n✓ Adding note computing formulas:")

# E34: Percentage of work done = (C21/C18)*100
ws.cell(34, 5).value = "% Work Done:"
ws.cell(34, 6).value = "=IF(C18,ROUND((C21/C18)*100,2),0)"
print(f"  F34: =IF(C18,ROUND((C21/C18)*100,2),0) → Percentage of work done")

# E35: Time allowed (days) = C14-C13
ws.cell(35, 5).value = "Time Allowed:"
ws.cell(35, 6).value = "=IF(AND(C14,C13),C14-C13,0)"
print(f"  F35: =IF(AND(C14,C13),C14-C13,0) → Time allowed in days")

# E36: Delay days = C15-C14
ws.cell(36, 5).value = "Delay Days:"
ws.cell(36, 6).value = "=IF(AND(C15,C14),C15-C14,0)"
print(f"  F36: =IF(AND(C15,C14),C15-C14,0) → Delay in days")

# E37: Extra item percentage = (C31/C18)*100
ws.cell(37, 5).value = "Extra Item %:"
ws.cell(37, 6).value = "=IF(AND(C31,C18),ROUND((C31/C18)*100,2),0)"
print(f"  F37: =IF(AND(C31,C18),ROUND((C31/C18)*100,2),0) → Extra item percentage")

# E38: Balance percentage = (C22/C18)*100
ws.cell(38, 5).value = "Balance %:"
ws.cell(38, 6).value = "=IF(C18,ROUND((C22/C18)*100,2),0)"
print(f"  F38: =IF(C18,ROUND((C22/C18)*100,2),0) → Balance percentage")

# E39: Delay as % of time allowed
ws.cell(39, 5).value = "Delay % of Time:"
ws.cell(39, 6).value = "=IF(F35,ROUND((F36/F35)*100,2),0)"
print(f"  F39: =IF(F35,ROUND((F36/F35)*100,2),0) → Delay as % of time allowed")

# Also add these as named ranges for easy reference
print("\n✓ Creating named ranges:")
try:
    wb.define_name('PercentageWorkDone', '=BillChecklist!$F$34')
    wb.define_name('TimeAllowedDays', '=BillChecklist!$F$35')
    wb.define_name('DelayDays', '=BillChecklist!$F$36')
    wb.define_name('ExtraItemPercentage', '=BillChecklist!$F$37')
    wb.define_name('BalancePercentage', '=BillChecklist!$F$38')
    wb.define_name('DelayPercentageOfTime', '=BillChecklist!$F$39')
    print("  ✓ Named ranges created")
except Exception as e:
    print(f"  ⚠ Could not create named ranges: {e}")

# Save the file
wb.save(source_file)
print(f"\n✓ Saved file with computing formulas")

# Verify
wb_check = openpyxl.load_workbook(source_file, data_only=False)
ws_check = wb_check.active

print("\n" + "="*70)
print("VERIFICATION - Computing Formulas:")
print("="*70)
for row in range(34, 40):
    label = ws_check.cell(row, 5).value
    formula = ws_check.cell(row, 6).value
    print(f"Row {row}: {label} = {formula}")

print("\n" + "="*70)
print("NOTE COMPUTING FORMULAS ADDED")
print("="*70)
print("\n✓ Computing formulas added in columns E-F (rows 34-39)")
print("✓ These formulas calculate:")
print("  • Percentage of work done")
print("  • Time allowed (days)")
print("  • Delay days")
print("  • Extra item percentage")
print("  • Balance percentage")
print("  • Delay as % of time allowed")
print("\n✓ VBA macro can reference these cells:")
print("  F34 = Percentage work done")
print("  F35 = Time allowed")
print("  F36 = Delay days")
print("  F37 = Extra item %")
print("  F38 = Balance %")
print("  F39 = Delay % of time")
print(f"\n✓ Backup saved: {backup_file}")
