"""
Analyze the workbook structure to understand the exact format
"""
import openpyxl
from openpyxl import load_workbook

workbook_path = 'ATTACHED_ASSETS/english Note FINAL BILL NOTE SHEET_UPDATED.xlsm'

print("="*70)
print("ANALYZING WORKBOOK STRUCTURE")
print("="*70)

try:
    wb = load_workbook(workbook_path, keep_vba=True, data_only=True)
    ws = wb.active
    
    print(f"\nActive Sheet: {ws.title}")
    print(f"Max Row: {ws.max_row}, Max Column: {ws.max_column}")
    
    # Read key cells based on VBA macro
    print("\n" + "="*70)
    print("KEY DATA CELLS (as per VBA macro):")
    print("="*70)
    
    key_cells = {
        'C5': 'Agreement Number',
        'C13': 'Date of Commencement',
        'C14': 'Schedule Completion',
        'C15': 'Actual Completion',
        'C18': 'Work Order Amount',
        'C19': '17.A - Upto Last Bill',
        'C20': '17.B - This Bill',
        'C21': '17.C - Total',
        'C29': 'Repair/Maintenance Work',
        'C30': 'Extra Item (Yes/No)',
        'C31': 'Extra Item Amount',
        'C32': 'Excess Quantity',
        'C33': 'Delay Comment',
        'A42': 'Output Note (merged A42:D42)'
    }
    
    for cell_ref, description in key_cells.items():
        cell = ws[cell_ref]
        value = cell.value
        if value is None:
            value = "(empty)"
        elif isinstance(value, str) and len(value) > 50:
            value = value[:50] + "..."
        print(f"{cell_ref:6} | {description:35} | {str(value)}")
    
    # Read row structure around key areas
    print("\n" + "="*70)
    print("ROW STRUCTURE (Rows 1-45):")
    print("="*70)
    
    for row in range(1, min(46, ws.max_row + 1)):
        row_data = []
        for col in ['A', 'B', 'C', 'D']:
            cell = ws[f"{col}{row}"]
            value = cell.value
            if value is None:
                value = ""
            elif isinstance(value, str):
                if len(value) > 30:
                    value = value[:30] + "..."
            else:
                value = str(value)
            row_data.append(f"{col}:{value}")
        
        # Only print rows with some content
        if any(cell.value for cell in [ws[f"A{row}"], ws[f"B{row}"], ws[f"C{row}"], ws[f"D{row}"]]):
            print(f"Row {row:2} | {' | '.join(row_data)}")
    
    # Check merged cells
    print("\n" + "="*70)
    print("MERGED CELLS:")
    print("="*70)
    for merged_range in ws.merged_cells.ranges:
        print(f"  {merged_range}")
    
    # Check column widths
    print("\n" + "="*70)
    print("COLUMN WIDTHS:")
    print("="*70)
    for col in ['A', 'B', 'C', 'D']:
        width = ws.column_dimensions[col].width
        print(f"  Column {col}: {width if width else 'default'}")
    
    wb.close()
    
except Exception as e:
    print(f"\nError: {e}")
    import traceback
    traceback.print_exc()

