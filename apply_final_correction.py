"""
Apply final correction to the original english Note FINAL BILL NOTE SHEET.xlsm:
1. Update structure with 17.A, 17.B, 17.C
2. Remove row 24 (Net Amount)
3. Update all formulas
4. Apply corrected VBA macro
"""
import openpyxl
from copy import copy
import shutil
from datetime import datetime
import win32com.client
import os
import time

# Create backup
source_file = 'ATTACHED_ASSETS/english Note FINAL BILL NOTE SHEET.xlsm'
backup_file = f'ATTACHED_ASSETS/english Note FINAL BILL NOTE SHEET_backup_final_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsm'
shutil.copy2(source_file, backup_file)
print(f"✓ Backup created: {backup_file}")

# Load the XLSM file
wb = openpyxl.load_workbook(source_file, keep_vba=True)
ws = wb.active

print(f"✓ Loaded workbook: {wb.sheetnames}")
print(f"✓ Active sheet: {ws.title}")

# Find row 19 (current item 17)
row_17_index = None
for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=30, min_col=1, max_col=1), start=1):
    cell_value = row[0].value
    if cell_value == 17:
        row_17_index = idx
        print(f"✓ Found item 17 at row {idx}")
        break

if not row_17_index:
    print("✗ Could not find item 17")
    exit(1)

# Insert 2 new rows after row 17
ws.insert_rows(row_17_index + 1, 2)
print(f"✓ Inserted 2 new rows after row {row_17_index}")

# Update row 19 (17.A)
ws.cell(row_17_index, 1).value = "17.A"
ws.cell(row_17_index, 2).value = "Sum of payment upto last bill Rs."
current_val = ws.cell(row_17_index, 3).value
if current_val and isinstance(current_val, str) and '=' in str(current_val):
    ws.cell(row_17_index, 3).value = 0
print(f"✓ Updated row {row_17_index} to 17.A")

# Create row 20 (17.B)
ws.cell(row_17_index + 1, 1).value = "B."
ws.cell(row_17_index + 1, 2).value = "Amount of this bill Rs."
ws.cell(row_17_index + 1, 3).value = 0
for col in range(1, 5):
    source_cell = ws.cell(row_17_index, col)
    target_cell = ws.cell(row_17_index + 1, col)
    if source_cell.has_style:
        target_cell.font = copy(source_cell.font)
        target_cell.border = copy(source_cell.border)
        target_cell.fill = copy(source_cell.fill)
        target_cell.number_format = copy(source_cell.number_format)
        target_cell.protection = copy(source_cell.protection)
        target_cell.alignment = copy(source_cell.alignment)
print(f"✓ Created row {row_17_index + 1} as 17.B")

# Create row 21 (17.C)
ws.cell(row_17_index + 2, 1).value = "C."
ws.cell(row_17_index + 2, 2).value = "Actual expenditure upto this bill = (A + B) Rs."
ws.cell(row_17_index + 2, 3).value = f"=C{row_17_index}+C{row_17_index + 1}"
for col in range(1, 5):
    source_cell = ws.cell(row_17_index, col)
    target_cell = ws.cell(row_17_index + 2, col)
    if source_cell.has_style:
        target_cell.font = copy(source_cell.font)
        target_cell.border = copy(source_cell.border)
        target_cell.fill = copy(source_cell.fill)
        target_cell.number_format = copy(source_cell.number_format)
        target_cell.protection = copy(source_cell.protection)
        target_cell.alignment = copy(source_cell.alignment)
print(f"✓ Created row {row_17_index + 2} as 17.C with formula =C{row_17_index}+C{row_17_index + 1}")

# Update formulas
import re

# Update Balance formula (item 18)
for row_idx in range(row_17_index + 3, min(row_17_index + 10, ws.max_row + 1)):
    if ws.cell(row_idx, 1).value == 18:
        item_16_row = None
        for r in range(1, row_17_index):
            if ws.cell(r, 1).value == 16:
                item_16_row = r
                break
        if item_16_row:
            ws.cell(row_idx, 3).value = f"=C{item_16_row}-C{row_17_index + 2}"
            print(f"✓ Updated item 18 (Balance) formula: =C{item_16_row}-C{row_17_index + 2}")
        break

# Find and update row 23 (Sum of Extra Items) - should have formula in C23
for row_idx in range(row_17_index + 3, min(row_17_index + 8, ws.max_row + 1)):
    cell_b = ws.cell(row_idx, 2).value
    if cell_b and "Sum of Extra Items" in str(cell_b):
        # Update formula in C23
        ws.cell(row_idx, 3).value = f"=IF(D{row_idx},ROUND((D{row_idx}/C{item_16_row})*100,2),0)"
        print(f"✓ Updated row {row_idx} (Extra Items) formula: =IF(D{row_idx},ROUND((D{row_idx}/C{item_16_row})*100,2),0)")
        
        # Now find and DELETE the "Net Amount of This Bill" row (should be next row)
        next_row = row_idx + 1
        next_cell_b = ws.cell(next_row, 2).value
        if next_cell_b and "Net Amount of This Bill" in str(next_cell_b):
            print(f"✓ Deleting row {next_row} (Net Amount of This Bill Rs.)")
            ws.delete_rows(next_row, 1)
        break

# Save temp file
temp_file = 'ATTACHED_ASSETS/english Note FINAL BILL NOTE SHEET_TEMP_FINAL.xlsm'
wb.save(temp_file)
print(f"\n✓ Saved structure updates to: {temp_file}")

# Show updated structure
print("\n" + "="*70)
print("UPDATED STRUCTURE:")
print("="*70)
wb_check = openpyxl.load_workbook(temp_file, data_only=False)
ws_check = wb_check.active
for row_idx in range(row_17_index - 2, min(row_17_index + 10, ws_check.max_row + 1)):
    cell_b = ws_check.cell(row_idx, 2).value
    if cell_b and len(str(cell_b)) > 40:
        cell_b = str(cell_b)[:40] + "..."
    print(f"Row {row_idx}: {ws_check.cell(row_idx, 1).value} | {cell_b} | {ws_check.cell(row_idx, 3).value}")

print("\n" + "="*70)
print("APPLYING VBA MACRO")
print("="*70)

# Apply VBA macro
try:
    xl = win32com.client.Dispatch("Excel.Application")
    xl.Visible = False
    xl.DisplayAlerts = False
    
    wb_com = xl.Workbooks.Open(os.path.abspath(temp_file))
    
    with open('updated_macro_corrected.vba', 'r', encoding='utf-8') as f:
        vba_code = f.read()
    
    module_found = False
    for component in wb_com.VBProject.VBComponents:
        if component.Type == 1:
            code = component.CodeModule.Lines(1, component.CodeModule.CountOfLines)
            if "GenerateBillNotes" in code:
                print(f"✓ Found macro in module: {component.Name}")
                component.CodeModule.DeleteLines(1, component.CodeModule.CountOfLines)
                component.CodeModule.AddFromString(vba_code)
                print(f"✓ Updated VBA code in module: {component.Name}")
                module_found = True
                break
    
    if not module_found:
        new_module = wb_com.VBProject.VBComponents.Add(1)
        new_module.Name = "BillNotesGenerator"
        new_module.CodeModule.AddFromString(vba_code)
        print(f"✓ Created new VBA module: BillNotesGenerator")
    
    # Save as final file
    final_file = os.path.abspath(source_file)
    wb_com.SaveAs(final_file, FileFormat=52)
    wb_com.Close(False)
    xl.Quit()
    
    print(f"✓ VBA macro applied successfully")
    
    # Clean up temp file
    time.sleep(1)
    if os.path.exists(temp_file):
        os.remove(temp_file)
        print(f"✓ Cleaned up temp file")
    
    print("\n" + "="*70)
    print("✅ FINAL CORRECTION COMPLETE!")
    print("="*70)
    print(f"✓ File updated: {source_file}")
    print(f"✓ Backup saved: {backup_file}")
    print("\nChanges applied:")
    print("  • Added rows 17.A, 17.B, 17.C")
    print("  • Removed 'Net Amount of This Bill' row")
    print("  • Updated all formulas")
    print("  • Applied corrected VBA macro")
    print("  • Output cell: B43")
    
except Exception as e:
    print(f"\n✗ VBA Error: {e}")
    print(f"\nStructure updated in: {temp_file}")
    print("Please manually copy VBA code from: updated_macro_corrected.vba")
